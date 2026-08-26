import json
import os
import requests
import tempfile
import threading
import unittest
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

os.environ["APP_SECRET_KEY"] = "test-secret"
_tmp_root = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
_tmp_path = Path(_tmp_root.name)
os.environ["DATA_DIR"] = str(_tmp_path / "data")
os.environ["DATABASE_PATH"] = str(_tmp_path / "data" / "test.db")
os.environ["UPLOADS_PRODUCTOS_DIR"] = str(_tmp_path / "uploads")

import app as cotizador_app
from app import (
    Cliente,
    Cotizacion,
    FamiliaCotizacion,
    Usuario,
    app,
    construir_link_cotizacion,
    db,
    generar_excel_cotizacion,
)
from openpyxl import load_workbook


class CotizadorCalculosTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        app.config["TESTING"] = True
        cls.tmp_root = _tmp_root

    @classmethod
    def tearDownClass(cls):
        with app.app_context():
            db.session.remove()
            db.engine.dispose()
        cls.tmp_root.cleanup()

    def setUp(self):
        with cotizador_app._bna_exchange_rate_lock:
            cotizador_app._bna_exchange_rate_cache["payload"] = None
            cotizador_app._bna_exchange_rate_cache["fetched_at"] = None
        with cotizador_app._bna_refresh_lock:
            cotizador_app._bna_refresh_in_progress = False
            cotizador_app._bna_refresh_last_error = None
        with app.app_context():
            db.session.remove()
            db.drop_all()
            db.create_all()

            usuario = Usuario(username="admin", nombre_completo="Administrador Principal", is_admin=True)
            usuario.set_password("admin123")
            operador = Usuario(username="operador", nombre_completo="Lucas Santacruz", is_admin=False)
            operador.set_password("operador123")
            cliente_exento = Cliente(
                nombre="Municipalidad de Lujan",
                razon_social="Municipalidad de Lujan",
                cuit="30-12345678-9",
                condicion_iva="Exento",
            )
            cliente_ri = Cliente(
                nombre="Cliente Responsable",
                razon_social="Cliente Responsable SA",
                cuit="30-87654321-0",
                condicion_iva="Responsable Inscrito",
            )
            familia = FamiliaCotizacion(nombre="SEGURIDAD URBANA", activa=True)
            db.session.add_all([usuario, operador, cliente_exento, cliente_ri, familia])
            db.session.commit()

            self.cliente_exento_id = cliente_exento.id
            self.cliente_ri_id = cliente_ri.id

        self.client = app.test_client()
        response = self.client.post(
            "/login",
            data={"username": "admin", "password": "admin123"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

    def _login(self, username, password):
        self.client.get("/logout")
        response = self.client.post(
            "/login",
            data={"username": username, "password": password},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

    def test_login_no_usa_llave_de_acceso_inicial(self):
        self.client.get("/logout")

        setup_response = self.client.get("/setup-admin")
        self.assertEqual(setup_response.status_code, 404)

        login_response = self.client.get("/login")
        self.assertEqual(login_response.status_code, 200)
        html = login_response.get_data(as_text=True)
        self.assertIn("Acceso al cotizador", html)
        self.assertIn("Usuario", html)
        self.assertIn('name="password"', html)
        self.assertNotIn("setup-admin", html)
        self.assertNotIn("ADMIN_SETUP_TOKEN", html)
        self.assertNotIn("Clave de instal", html)

    def test_login_prioriza_usuario_con_mayusculas_exactas(self):
        with app.app_context():
            usuario = Usuario(username="Admin", nombre_completo="Administrador Alternativo", is_admin=True)
            usuario.set_password("otra-clave")
            db.session.add(usuario)
            db.session.commit()

        self.client.get("/logout")
        response = self.client.post(
            "/login",
            data={"username": "admin", "password": "admin123"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers.get("Location"), "/")

    def _form_cotizacion(self, **overrides):
        data = {
            "cliente_id": str(self.cliente_exento_id),
            "cliente": "Municipalidad de Lujan",
            "cliente_contacto": "Martin Moreno",
            "cliente_razon_social": "Municipalidad de Lujan",
            "cliente_cuit": "30-12345678-9",
            "familia": "SEGURIDAD URBANA",
            "moneda": "USD",
            "tipo_cambio": "1450",
            "condicion_iva": "Exento",
            "estado": "En progreso",
            "bonificacion_cierre_monto": "0",
            "seguimiento_email": "",
            "seguimiento_cada_dias": "7",
            "row_id[]": ["row-1"],
            "item_id[]": [""],
            "imagen_actual[]": [""],
            "desc[]": ["Lector DS-K1T321MFWX"],
            "detalle[]": ["Control de acceso Hikvision con reconocimiento facial"],
            "cant[]": ["1"],
            "costo[]": ["95"],
            "iva_compra[]": ["21"],
            "extra[]": ["5"],
            "margen[]": ["60"],
            "descuento[]": ["0"],
            "carga_fiscal[]": ["5"],
            "iva_item[]": ["21"],
        }
        data.update(overrides)
        return data

    def _crear_cotizacion(self, **overrides):
        response = self.client.post("/", data=self._form_cotizacion(**overrides), follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            cotizacion = Cotizacion.query.order_by(Cotizacion.id.desc()).first()
            self.assertIsNotNone(cotizacion)
            return cotizacion.id

    def _cotizacion(self, cotizacion_id):
        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            if cotizacion:
                list(cotizacion.items)
            return cotizacion

    def _set_fecha_cotizacion(self, cotizacion_id, fecha):
        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            cotizacion.fecha = fecha
            db.session.commit()

    def _valor_por_etiqueta(self, ws, etiqueta):
        for row in ws.iter_rows():
            for indice, cell in enumerate(row[:-1]):
                if cell.value == etiqueta:
                    return row[indice + 1].value
        self.fail(f"No se encontro la etiqueta {etiqueta!r} en el Excel")

    def test_calculo_completo_y_excel_quedan_alineados(self):
        cotizacion_id = self._crear_cotizacion()

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            item = cotizacion.items[0]

            self.assertEqual(cotizacion.moneda, "USD")
            self.assertAlmostEqual(cotizacion.tipo_cambio_usado, 1450.0, places=4)
            self.assertAlmostEqual(item.precio_venta, 159.60, places=2)
            self.assertAlmostEqual(cotizacion.total_neto, 159.60, places=2)
            self.assertAlmostEqual(cotizacion.total_iva, 33.52, places=2)
            self.assertAlmostEqual(cotizacion.total_final, 193.12, places=2)
            self.assertAlmostEqual(cotizacion.total_carga_fiscal, 7.98, places=2)

            wb = load_workbook(BytesIO(generar_excel_cotizacion(cotizacion)), data_only=True)

        ws = wb.active
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Tipo de cambio usado"), 1450.0, places=4)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "IVA compra credito"), 19.95, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "IVA a pagar estimado"), 13.57, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Carga fiscal retenida"), 7.98, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Ganancia neta (Bolsillo)"), 51.87, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Total a cobrar"), 193.12, places=2)

    def test_formulario_muestra_ayudas_de_calculo(self):
        payload_bna = {
            "ok": True,
            "rate": 1450.0,
            "date": "24/04/2026",
            "time": "12:00",
            "label": "BNA billete vendedor",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna):
            response = self.client.get("/")

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Criterio interno de calculo", html)
        self.assertIn('data-bs-toggle="tooltip"', html)
        self.assertIn("Costo Neto U.", html)
        self.assertIn("IVA Compra %", html)
        self.assertIn("Bonificacion de cierre", html)
        self.assertIn("Ganancia Neta (Bolsillo)", html)
        self.assertIn("inicializarAyudasCalculo", html)
        self.assertIn("Una condicion por linea.", html)
        self.assertIn("Las condiciones base ya vienen cargadas.", html)
        self.assertIn("Las imagenes son ilustrativas.", html)
        self.assertNotIn("Agregar personalizada", html)

    def test_formulario_cliente_tiene_guardado_con_manejo_de_errores(self):
        payload_bna = {
            "ok": True,
            "rate": 1450.0,
            "date": "24/04/2026",
            "time": "12:00",
            "label": "BNA billete vendedor",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna):
            response = self.client.get("/")

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("async function guardarCliente()", html)
        self.assertIn("Guardando cambios...", html)
        self.assertIn(".catch(() => ({}))", html)
        self.assertIn("No se pudo guardar el cliente", html)
        self.assertIn("modal-dialog-scrollable", html)
        self.assertIn("function moverModalClienteAlBody()", html)
        self.assertIn("El servidor no devolvio los datos del cliente guardado", html)
        self.assertIn("function actualizarSubsectores(valorSeleccionado = \"\")", html)

    def test_formulario_tiene_borrador_automatico_para_no_perder_carga(self):
        payload_bna = {
            "ok": True,
            "rate": 1450.0,
            "date": "24/04/2026",
            "time": "12:00",
            "label": "BNA billete vendedor",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna):
            response = self.client.get("/")

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Borrador recuperado", html)
        self.assertIn("cuencotech:cotizacion:draft:new", html)
        self.assertIn("function inicializarBorradorCotizacion()", html)
        self.assertIn("function aplicarBorradorCotizacion(draft)", html)
        self.assertIn("function marcarBorradorComoPendienteDeEnvio()", html)
        self.assertIn("bonificacion_cierre_monto", html)
        self.assertIn("localStorage.setItem(cotizadorDraftKey", html)
        self.assertIn("validarImagenesAntesDeEnviar", html)
        self.assertIn("limpiarBorradorCotizadorGuardado", html)

    def test_frontend_normaliza_error_ssl_del_bna(self):
        payload_bna = {
            "ok": False,
            "rate": None,
            "date": "",
            "time": "",
            "label": "BNA billete vendedor",
            "error": "HTTPSConnectionPool(host='www.bna.com.ar', port=443): Max retries exceeded with url: /Personas (Caused by SSLError(SSLCertVerificationError(1, '[SSL: CERTIFICATE_VERIFY_FAILED] certificate verify failed: self-signed certificate in certificate chain (_ssl.c:1016)')))",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna):
            response = self.client.get("/")

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("function normalizarMensajeErrorTipoCambio", html)
        self.assertIn("No se pudo validar la conexion segura con el BNA.", html)

    def test_renderizar_cotizador_no_consulta_bna(self):
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna") as obtener_bna:
            response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        obtener_bna.assert_not_called()
        self.assertIn("El tipo de cambio se cargara desde BNA al abrir el cotizador.", response.get_data(as_text=True))

    def test_endpoint_publica_metricas_de_duracion_y_consultas_sql(self):
        with patch.object(app.logger, "info") as logger_info:
            response = self.client.get("/historial")

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(float(response.headers["X-Response-Time-Ms"]), 0.0)
        self.assertGreaterEqual(int(response.headers["X-SQL-Query-Count"]), 1)
        server_timing = response.headers["Server-Timing"]
        self.assertIn("app;dur=", server_timing)
        self.assertIn("sql;dur=", server_timing)
        self.assertIn("queries", server_timing)
        self.assertTrue(
            any(
                llamada.args and llamada.args[0].startswith("request_metrics endpoint=")
                for llamada in logger_info.call_args_list
            )
        )

    def test_cotizacion_y_auditoria_se_confirman_en_una_sola_transaccion(self):
        commits = []

        def registrar_commit(_connection):
            commits.append(1)

        with app.app_context():
            engine = db.engine
            cotizador_app.event.listen(engine, "commit", registrar_commit)
        try:
            response = self.client.post(
                "/",
                data=self._form_cotizacion(),
                follow_redirects=False,
            )
        finally:
            cotizador_app.event.remove(engine, "commit", registrar_commit)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(commits), 1)
        with app.app_context():
            cotizacion = Cotizacion.query.order_by(Cotizacion.id.desc()).first()
            auditorias = cotizador_app.Auditoria.query.filter_by(
                tipo_entidad="Cotización",
                entidad_id=cotizacion.id,
            ).all()
            self.assertEqual(len(auditorias), 1)

    def test_filtrar_historial_evita_consultas_n_mas_uno_de_clientes(self):
        with app.app_context():
            ahora = cotizador_app.datetime.utcnow()
            db.session.add_all(
                [
                    Cotizacion(
                        numero_cotizacion=f"CT-TEST-{indice:03d}",
                        estado="En progreso",
                        cliente=f"Cliente {indice}",
                        cliente_id=self.cliente_exento_id,
                        familia="SEGURIDAD URBANA",
                        fecha=ahora,
                        moneda="USD",
                        total_final=float(indice),
                    )
                    for indice in range(25)
                ]
            )
            db.session.commit()

        response = self.client.get("/filtrar_historial?page=1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.get_json()["items"]), cotizador_app.HISTORIAL_PER_PAGE)
        self.assertLessEqual(int(response.headers["X-SQL-Query-Count"]), 4)
        self.assertTrue(
            all(
                item["cliente_nombre"] == "Municipalidad de Lujan"
                for item in response.get_json()["items"]
            )
        )

    def test_guardados_concurrentes_generan_numeros_unicos(self):
        workers = 6
        barrera = threading.Barrier(workers)
        with app.app_context():
            usuario = Usuario.query.filter_by(username="admin").one()
            token = cotizador_app.generar_token_usuario(usuario)

        def guardar_cotizacion(_indice):
            cliente = app.test_client()
            cliente.set_cookie("x-access-token", token)
            barrera.wait()
            respuesta = cliente.post(
                "/",
                data=self._form_cotizacion(),
                follow_redirects=False,
            )
            return respuesta.status_code

        with ThreadPoolExecutor(max_workers=workers) as pool:
            estados = list(pool.map(guardar_cotizacion, range(workers)))

        self.assertEqual(estados, [302] * workers)
        with app.app_context():
            numeros = [fila[0] for fila in db.session.query(Cotizacion.numero_cotizacion).all()]
            auditorias = cotizador_app.Auditoria.query.filter_by(tipo_entidad="Cotización").count()
        self.assertEqual(len(numeros), workers)
        self.assertEqual(len(set(numeros)), workers)
        self.assertEqual(auditorias, workers)

    def test_tipo_cambio_bna_cache_persiste_entre_memorias(self):
        payload = {
            "ok": True,
            "rate": 1450.0,
            "buy_rate": 1400.0,
            "label": "BNA billete vendedor",
        }
        with app.app_context():
            cotizador_app.guardar_tipo_cambio_bna_cache(payload)
            with cotizador_app._bna_exchange_rate_lock:
                cotizador_app._bna_exchange_rate_cache["payload"] = None
                cotizador_app._bna_exchange_rate_cache["fetched_at"] = None
            recuperado = cotizador_app.obtener_tipo_cambio_bna_cache()

        self.assertEqual(recuperado["rate"], 1450.0)
        self.assertFalse(recuperado["stale"])

    def test_renderizar_cotizador_recupera_cache_persistente_sin_consultar_bna(self):
        payload = {"ok": True, "rate": 1450.0, "label": "BNA billete vendedor"}
        with app.app_context():
            cotizador_app.guardar_tipo_cambio_bna_cache(payload)
            with cotizador_app._bna_exchange_rate_lock:
                cotizador_app._bna_exchange_rate_cache["payload"] = None
                cotizador_app._bna_exchange_rate_cache["fetched_at"] = None

        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna") as obtener_bna:
            response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertIn('value="1450.0"', response.get_data(as_text=True))
        obtener_bna.assert_not_called()

    def test_api_tipo_cambio_responde_cache_y_dispara_actualizacion(self):
        payload = {"ok": True, "rate": 1450.0, "label": "BNA billete vendedor"}
        with app.app_context():
            cotizador_app.guardar_tipo_cambio_bna_cache(payload)
        with patch.object(cotizador_app, "solicitar_actualizacion_tipo_cambio_bna", return_value=True) as solicitar:
            with patch.object(cotizador_app, "estado_actualizacion_tipo_cambio_bna", return_value=(True, None)):
                response = self.client.get("/api/tipo-cambio/oficial?refresh=1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["rate"], 1450.0)
        self.assertTrue(response.get_json()["refreshing"])
        solicitar.assert_called_once_with(force=True)

    def test_actualizacion_bna_en_segundo_plano_fuerza_refresco(self):
        payload = {"ok": True, "rate": 1450.0}
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload) as obtener_bna:
            cotizador_app._actualizar_tipo_cambio_bna_en_segundo_plano(force=False)

        obtener_bna.assert_called_once_with(force=True)

    def test_tipo_cambio_vencido_se_devuelve_sin_consultar_bna(self):
        payload = {"ok": True, "rate": 1450.0, "label": "BNA billete vendedor"}
        with app.app_context():
            cotizador_app.guardar_tipo_cambio_bna_cache(payload)
            cache_persistente = db.session.get(cotizador_app.TipoCambioBnaCache, 1)
            cache_persistente.actualizado_en = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(
                seconds=cotizador_app.BNA_TC_CACHE_SECONDS + 1
            )
            db.session.commit()
            with cotizador_app._bna_exchange_rate_lock:
                cotizador_app._bna_exchange_rate_cache["payload"] = None
                cotizador_app._bna_exchange_rate_cache["fetched_at"] = None
            with patch.object(cotizador_app, "descargar_html_bna") as descargar_bna:
                recuperado = cotizador_app.obtener_tipo_cambio_oficial_bna()

        self.assertEqual(recuperado["rate"], 1450.0)
        self.assertTrue(recuperado["stale"])
        descargar_bna.assert_not_called()

    def test_descarga_bna_usa_timeout_corto(self):
        class FakeResponse:
            text = "<html></html>"

            def raise_for_status(self):
                return None

        with patch.dict(os.environ, {"BNA_SSL_MODE": "strict"}):
            with patch.object(cotizador_app.requests, "get", return_value=FakeResponse()) as mock_get:
                html, ssl_insecure = cotizador_app.descargar_html_bna()

        self.assertEqual(html, "<html></html>")
        self.assertFalse(ssl_insecure)
        self.assertEqual(cotizador_app.BNA_REQUEST_TIMEOUT_SECONDS, 3)
        self.assertEqual(mock_get.call_args.kwargs["timeout"], 3)

    def test_tipo_cambio_bna_reintenta_sin_validacion_ssl_en_modo_auto(self):
        payload_bna = {
            "ok": True,
            "rate": 1385.0,
            "buy_rate": 1335.0,
            "date": "04/05/2026",
            "time": "13:00",
            "label": "BNA billete vendedor",
            "source_code": "billetes_venta",
            "item": "Dolar U.S.A",
            "source_url": cotizador_app.BNA_PERSONAS_URL,
        }

        class FakeResponse:
            text = "<html></html>"

            def raise_for_status(self):
                return None

        with patch.dict(os.environ, {"BNA_SSL_MODE": "auto"}):
            with patch.object(
                cotizador_app.requests,
                "get",
                side_effect=[requests.exceptions.SSLError("cert"), FakeResponse()],
            ) as mock_get:
                with patch.object(cotizador_app, "extraer_tipo_cambio_bna", return_value=payload_bna):
                    resultado = cotizador_app.obtener_tipo_cambio_oficial_bna(force=True)

        self.assertTrue(resultado["ok"])
        self.assertTrue(resultado["ssl_insecure"])
        self.assertEqual(mock_get.call_count, 2)
        self.assertTrue(mock_get.call_args_list[0].kwargs["verify"])
        self.assertFalse(mock_get.call_args_list[1].kwargs["verify"])

    def test_api_actualizar_cliente_responde_json_y_actualiza_datos(self):
        response = self.client.put(
            f"/api/clientes/{self.cliente_exento_id}",
            json={
                "nombre": "Municipalidad de Lujan Editada",
                "razon_social": "Municipalidad de Lujan",
                "cuit": "30-12345678-9",
                "domicilio": "San Martin 100",
                "sector": "Publico",
                "subsector": "Municipal",
                "email": "test@lujan.gob.ar",
                "telefono": "123",
                "condicion_iva": "Exento",
            },
        )

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["nombre"], "Municipalidad de Lujan Editada")
        self.assertEqual(data["sector"], "Publico")
        self.assertEqual(data["subsector"], "Municipal")
        with app.app_context():
            cliente = db.session.get(Cliente, self.cliente_exento_id)
            self.assertEqual(cliente.nombre, "Municipalidad de Lujan Editada")

    def test_api_actualizar_cliente_sin_sesion_responde_json_auth_required(self):
        self.client.get("/logout")
        response = self.client.put(
            f"/api/clientes/{self.cliente_exento_id}",
            json={"nombre": "Sin sesion", "sector": "Publico", "subsector": "Municipal"},
        )

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.get_json()["error"], "auth_required")

    def test_admin_puede_actualizar_nombre_visible_de_usuario(self):
        with app.app_context():
            operador = Usuario.query.filter_by(username="operador").first()
            operador_id = operador.id

        response = self.client.post(
            f"/usuarios/{operador_id}/nombre",
            data={"nombre_completo": "Lucas A. Santacruz"},
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        with app.app_context():
            operador = db.session.get(Usuario, operador_id)
            self.assertEqual(operador.nombre_completo, "Lucas A. Santacruz")

    def test_usuario_operador_no_ve_ni_accede_dashboard_o_auditoria(self):
        self._login("operador", "operador123")

        payload_bna = {
            "ok": True,
            "rate": 1450.0,
            "date": "24/04/2026",
            "time": "12:00",
            "label": "BNA billete vendedor",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna):
            response = self.client.get("/")

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Operador", html)
        self.assertIn('href="/familias"', html)
        self.assertNotIn('href="/dashboard"', html)
        self.assertNotIn('href="/auditoria"', html)
        self.assertNotIn('href="/usuarios"', html)

        for path in ("/dashboard", "/dashboard/detalle-operativo", "/auditoria"):
            response = self.client.get(path, follow_redirects=False)
            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/"))

    def test_usuario_operador_puede_gestionar_familias_y_editar_cotizaciones(self):
        cotizacion_id = self._crear_cotizacion()
        self._login("operador", "operador123")

        response = self.client.get("/familias")
        self.assertEqual(response.status_code, 200)
        self.assertIn("SEGURIDAD URBANA", response.get_data(as_text=True))

        response = self.client.post("/familias", data={"nombre": "DATA CENTER"}, follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            familia = FamiliaCotizacion.query.filter_by(nombre="DATA CENTER").first()
            self.assertIsNotNone(familia)
            familia_id = familia.id

        response = self.client.post(
            f"/familias/{familia_id}/editar",
            data={"nombre": "DATA CENTER PLUS", "activa": "1"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            self.assertIsNotNone(FamiliaCotizacion.query.filter_by(nombre="DATA CENTER PLUS", activa=True).first())

        response = self.client.post("/api/familias", json={"nombre": "INTEGRACIONES"})
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.get_json()["nombre"], "INTEGRACIONES")

        response = self.client.get(f"/cotizacion/{cotizacion_id}/editar")
        self.assertEqual(response.status_code, 200)

    def test_historial_permite_clonar_y_clonado_duplica_cotizacion(self):
        cotizacion_id = self._crear_cotizacion(
            condicion_iva="Responsable Inscrito",
            bonificacion_cierre_monto="50",
            **{"descuento[]": ["10"]},
        )

        with app.app_context():
            cotizacion_original = db.session.get(Cotizacion, cotizacion_id)
            cotizacion_original.estado = "Aceptada"
            cotizacion_original.seguimiento_activo = True
            cotizacion_original.seguimiento_email = "seguimiento@cuencotech.com"
            cotizacion_original.seguimiento_cada_dias = 5
            cotizacion_original.seguimiento_proximo_envio = cotizador_app.datetime.utcnow()
            cotizacion_original.seguimiento_ultimo_envio = cotizador_app.datetime.utcnow()
            item_original = cotizacion_original.items[0]
            imagen_relativa = "uploads/productos/test-clone-source.jpg"
            (Path(app.static_folder) / "uploads" / "productos").mkdir(parents=True, exist_ok=True)
            (Path(app.static_folder) / imagen_relativa).write_bytes(b"imagen-clon")
            item_original.imagen_url = imagen_relativa
            db.session.commit()

        response_historial = self.client.get("/historial")
        html_historial = response_historial.get_data(as_text=True)
        self.assertEqual(response_historial.status_code, 200)
        self.assertIn("function clonarCotizacion", html_historial)
        self.assertIn("bi-copy", html_historial)
        self.assertIn("/cotizacion/clonar/", html_historial)

        response = self.client.post(f"/cotizacion/clonar/{cotizacion_id}")
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertTrue(data["ok"])
        self.assertIn("/editar", data["redirect_url"])

        with app.app_context():
            cotizacion_original = db.session.get(Cotizacion, cotizacion_id)
            cotizacion_clonada = db.session.get(Cotizacion, data["nueva_id"])
            self.assertIsNotNone(cotizacion_clonada)
            self.assertNotEqual(cotizacion_clonada.id, cotizacion_original.id)
            self.assertNotEqual(cotizacion_clonada.numero_cotizacion, cotizacion_original.numero_cotizacion)
            self.assertEqual(cotizacion_clonada.estado, "En progreso")
            self.assertGreaterEqual(cotizacion_clonada.fecha, cotizacion_original.fecha)
            self.assertEqual(cotizacion_clonada.cliente, cotizacion_original.cliente)
            self.assertEqual(cotizacion_clonada.cliente_id, cotizacion_original.cliente_id)
            self.assertEqual(cotizacion_clonada.cliente_contacto, cotizacion_original.cliente_contacto)
            self.assertEqual(cotizacion_clonada.familia, cotizacion_original.familia)
            self.assertEqual(cotizacion_clonada.condicion_iva, cotizacion_original.condicion_iva)
            self.assertEqual(cotizacion_clonada.forma_pago, cotizacion_original.forma_pago)
            self.assertEqual(cotizacion_clonada.condicion_cotizacion, cotizacion_original.condicion_cotizacion)
            self.assertEqual(cotizacion_clonada.observacion_cliente, cotizacion_original.observacion_cliente)
            self.assertAlmostEqual(cotizacion_clonada.total_neto, cotizacion_original.total_neto, places=2)
            self.assertAlmostEqual(cotizacion_clonada.total_iva, cotizacion_original.total_iva, places=2)
            self.assertAlmostEqual(cotizacion_clonada.total_final, cotizacion_original.total_final, places=2)
            self.assertAlmostEqual(cotizacion_clonada.total_carga_fiscal, cotizacion_original.total_carga_fiscal, places=2)
            self.assertAlmostEqual(
                cotizacion_clonada.bonificacion_cierre_monto,
                cotizacion_original.bonificacion_cierre_monto,
                places=2,
            )
            self.assertFalse(cotizacion_clonada.seguimiento_activo)
            self.assertIsNone(cotizacion_clonada.seguimiento_proximo_envio)
            self.assertIsNone(cotizacion_clonada.seguimiento_ultimo_envio)
            self.assertEqual(len(cotizacion_clonada.items), len(cotizacion_original.items))

            item_clonado = cotizacion_clonada.items[0]
            item_original = cotizacion_original.items[0]
            self.assertEqual(item_clonado.descripcion, item_original.descripcion)
            self.assertEqual(item_clonado.detalle, item_original.detalle)
            self.assertEqual(item_clonado.cantidad, item_original.cantidad)
            self.assertAlmostEqual(item_clonado.costo_unitario, item_original.costo_unitario, places=2)
            self.assertAlmostEqual(item_clonado.iva_compra_pct, item_original.iva_compra_pct, places=2)
            self.assertAlmostEqual(item_clonado.costo_extra, item_original.costo_extra, places=2)
            self.assertAlmostEqual(item_clonado.margen, item_original.margen, places=4)
            self.assertAlmostEqual(item_clonado.descuento_pct, item_original.descuento_pct, places=2)
            self.assertAlmostEqual(item_clonado.carga_fiscal, item_original.carga_fiscal, places=2)
            self.assertAlmostEqual(item_clonado.iva_item, item_original.iva_item, places=2)
            self.assertAlmostEqual(item_clonado.precio_venta, item_original.precio_venta, places=2)
            self.assertAlmostEqual(item_clonado.subtotal, item_original.subtotal, places=2)
            self.assertNotEqual(item_clonado.imagen_url, item_original.imagen_url)
            self.assertTrue((Path(app.static_folder) / item_original.imagen_url).exists())
            self.assertTrue((cotizador_app.UPLOADS_PRODUCTOS_DIR / Path(item_clonado.imagen_url).name).exists())

    def test_clonar_cotizacion_sin_sesion_responde_json_auth_required(self):
        cotizacion_id = self._crear_cotizacion()
        self.client.get("/logout")

        response = self.client.post(f"/cotizacion/clonar/{cotizacion_id}")

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.get_json()["error"], "auth_required")

    def test_detalle_operativo_dashboard_filtra_por_familia_propia(self):
        seguridad_id = self._crear_cotizacion(
            cliente_id="",
            cliente="Cliente Seguridad",
            cliente_razon_social="Cliente Seguridad SA",
            cliente_cuit="30-11111111-1",
            familia="SEGURIDAD URBANA",
        )
        with app.app_context():
            db.session.add(FamiliaCotizacion(nombre="DATA CENTER", activa=True))
            db.session.commit()
        data_center_id = self._crear_cotizacion(
            cliente_id="",
            cliente="Cliente Data Center",
            cliente_razon_social="Cliente Data Center SA",
            cliente_cuit="30-22222222-2",
            familia="DATA CENTER",
        )
        fecha_dentro_del_rango = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(days=1)
        self._set_fecha_cotizacion(seguridad_id, fecha_dentro_del_rango)
        self._set_fecha_cotizacion(data_center_id, fecha_dentro_del_rango)

        response = self.client.get("/dashboard/detalle-operativo?periodo=365&op_familia=DATA%20CENTER")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('name="op_familia"', html)
        self.assertIn("Familia del detalle", html)
        self.assertIn("Cliente Data Center", html)
        self.assertNotIn("Cliente Seguridad", html)

    def test_dashboard_ignora_ars_y_muestra_solo_usd(self):
        cotizacion_usd_id = self._crear_cotizacion(
            cliente_id="",
            cliente="Cliente USD",
            cliente_razon_social="Cliente USD SA",
            cliente_cuit="30-33333333-3",
            moneda="USD",
            familia="SEGURIDAD URBANA",
        )
        cotizacion_ars_id = self._crear_cotizacion(
            cliente_id="",
            cliente="Cliente ARS",
            cliente_razon_social="Cliente ARS SA",
            cliente_cuit="30-44444444-4",
            moneda="ARS",
            tipo_cambio="1",
            familia="SEGURIDAD URBANA",
        )
        fecha_dentro_del_rango = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(days=1)
        self._set_fecha_cotizacion(cotizacion_usd_id, fecha_dentro_del_rango)
        self._set_fecha_cotizacion(cotizacion_ars_id, fecha_dentro_del_rango)

        response = self.client.get("/dashboard?periodo=365&moneda=ARS")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('name="moneda"', html)
        self.assertNotIn("Pesos argentinos", html)
        self.assertNotIn("Total ARS", html)
        self.assertIn("Volumen cotizado USD", html)
        self.assertIn("Cliente USD", html)
        self.assertNotIn("Cliente ARS", html)

        response = self.client.get("/dashboard/detalle-operativo?periodo=365&moneda=ARS")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Total USD", html)
        self.assertIn("Cliente USD", html)
        self.assertNotIn("Cliente ARS", html)

    def test_dashboard_calcula_resumen_series_y_desgloses_con_sql(self):
        with app.app_context():
            db.session.add(FamiliaCotizacion(nombre="DATA CENTER", activa=True))
            cliente = db.session.get(Cliente, self.cliente_exento_id)
            cliente.sector = "Publico"
            cliente.subsector = "Municipal"
            db.session.commit()

        cotizacion_aceptada_id = self._crear_cotizacion(familia="SEGURIDAD URBANA")
        cotizacion_rechazada_id = self._crear_cotizacion(familia="DATA CENTER")
        cotizacion_pipeline_id = self._crear_cotizacion(
            cliente_id="",
            cliente="Cliente sin ficha",
            cliente_razon_social="Cliente sin ficha SA",
            cliente_cuit="30-55555555-5",
            familia="DATA CENTER",
        )
        fecha = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(days=1)

        with app.app_context():
            cotizacion_aceptada = db.session.get(Cotizacion, cotizacion_aceptada_id)
            cotizacion_aceptada.estado = "Aceptada"
            cotizacion_aceptada.total_final = 100.0
            cotizacion_aceptada.fecha = fecha

            cotizacion_rechazada = db.session.get(Cotizacion, cotizacion_rechazada_id)
            cotizacion_rechazada.estado = "Rechazada"
            cotizacion_rechazada.total_final = 250.0
            cotizacion_rechazada.fecha = fecha

            cotizacion_pipeline = db.session.get(Cotizacion, cotizacion_pipeline_id)
            cotizacion_pipeline.estado = "En progreso"
            cotizacion_pipeline.total_final = 50.0
            cotizacion_pipeline.fecha = fecha
            db.session.commit()

            desde_date = fecha.date() - cotizador_app.timedelta(days=1)
            hasta_date = fecha.date() + cotizador_app.timedelta(days=1)
            filtros = {
                "desde": desde_date.isoformat(),
                "hasta": hasta_date.isoformat(),
                "cliente": "",
                "familia": "",
                "sector": "",
                "subsector": "",
                "moneda": "USD",
            }
            query = cotizador_app.construir_query_dashboard_periodo(filtros)
            resumen = cotizador_app.construir_resumen_dashboard_sql(query)
            series = cotizador_app.construir_series_dashboard_sql(query, desde_date, hasta_date)
            clientes = cotizador_app.construir_top_clientes_dashboard_sql(query)
            familias = cotizador_app.construir_desglose_dashboard_sql(
                query,
                Cotizacion.familia,
                resumen["total"],
                default_label="Sin familia",
            )
            sectores = cotizador_app.construir_desglose_dashboard_sql(
                query,
                Cliente.sector,
                resumen["total"],
                default_label="Sin sector",
            )
            agrupaciones = cotizador_app.construir_agrupaciones_dashboard_sql(query, resumen["total"])

        self.assertEqual(resumen["total"], 3)
        self.assertEqual(resumen["aceptadas"], 1)
        self.assertEqual(resumen["rechazadas"], 1)
        self.assertEqual(resumen["en_progreso"], 1)
        self.assertEqual(resumen["total_importe"], 400.0)
        self.assertEqual(resumen["importe_pipeline"], 50.0)
        self.assertEqual(resumen["importe_aceptado"], 100.0)
        self.assertEqual(sum(series["creadas"]), 3)
        self.assertEqual(sum(series["cerradas"]), 2)
        self.assertEqual(sum(series["aceptadas"]), 1)
        self.assertEqual(series["resumen"], resumen)
        self.assertEqual(clientes[0]["nombre"], "Municipalidad de Lujan")
        self.assertEqual(clientes[0]["cantidad"], 2)
        self.assertEqual(clientes[0]["total"], 350.0)
        self.assertEqual(familias[0]["nombre"], "DATA CENTER")
        self.assertEqual(familias[0]["cantidad"], 2)
        self.assertEqual(familias[0]["porcentaje"], 66.7)
        self.assertEqual(sectores[0]["nombre"], "Publico")
        self.assertEqual(sectores[0]["cantidad"], 2)
        self.assertEqual(sectores[1]["nombre"], "Sin sector")
        self.assertEqual(sectores[1]["cantidad"], 1)
        self.assertEqual(agrupaciones["clientes"], clientes)
        self.assertEqual(agrupaciones["familias"], familias)
        self.assertEqual(agrupaciones["sectores"], sectores)

    def test_links_de_recordatorio_usan_app_base_url_configurada(self):
        cotizacion_id = self._crear_cotizacion()
        cotizacion = self._cotizacion(cotizacion_id)

        with patch.dict(os.environ, {"APP_BASE_URL": "http://localhost:9000"}):
            self.assertEqual(
                construir_link_cotizacion(cotizacion, editar=False),
                f"http://localhost:9000/cotizacion/{cotizacion_id}",
            )
            self.assertEqual(
                construir_link_cotizacion(cotizacion, editar=True),
                f"http://localhost:9000/cotizacion/{cotizacion_id}/editar",
            )

    def test_cliente_exento_no_ve_iva_detallado_pero_el_sistema_lo_calcula(self):
        cotizacion_id = self._crear_cotizacion()

        response = self.client.get(f"/cotizacion/{cotizacion_id}")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("Tipo de cambio usado", html)
        self.assertIn("Los precios incluyen IVA", html)
        self.assertIn(">Precio</th>", html)
        self.assertNotIn(">P. unit.</th>", html)
        self.assertIn("Condiciones comerciales", html)
        self.assertIn("Emitido por", html)
        self.assertIn("Administrador Principal", html)
        self.assertNotIn("IVA Total:", html)

        cotizacion = self._cotizacion(cotizacion_id)
        self.assertAlmostEqual(cotizacion.total_iva, 33.52, places=2)
        self.assertEqual(cotizacion.condicion_iva, "Exento")

    def test_cotizacion_guarda_y_muestra_forma_pago_condicion_y_observacion(self):
        condiciones = [
            "Las imagenes son ilustrativas.",
            "Oferta sujeta a stock.",
            "Entrega y validez a confirmar.",
        ]
        cotizacion_id = self._crear_cotizacion(
            forma_pago="30 dias",
            condicion_cotizacion=json.dumps(condiciones, ensure_ascii=False),
            observacion_cliente="Solicitud de presupuesto n 4587 - Lic municipal 221304",
        )

        cotizacion = self._cotizacion(cotizacion_id)
        self.assertEqual(cotizacion.forma_pago, "30 dias")
        self.assertEqual(cotizacion.cliente_contacto, "Martin Moreno")
        self.assertEqual(cotizacion.condiciones_cotizacion_lista, condiciones)
        self.assertEqual(cotizacion.observacion_cliente, "Solicitud de presupuesto n 4587 - Lic municipal 221304")

        html = self.client.get(f"/cotizacion/{cotizacion_id}").get_data(as_text=True)
        self.assertIn("Forma de pago", html)
        self.assertIn("30 dias", html)
        self.assertIn("Solicitud de presupuesto n 4587 - Lic municipal 221304", html)
        for condicion in condiciones:
            self.assertIn(condicion, html)
        self.assertNotIn('["Las imagenes son ilustrativas."', html)

        with app.app_context():
            cotizacion_db = db.session.get(Cotizacion, cotizacion_id)
            wb = load_workbook(BytesIO(generar_excel_cotizacion(cotizacion_db)), data_only=True)

        ws = wb.active
        self.assertEqual(self._valor_por_etiqueta(ws, "Contacto cliente"), "Martin Moreno")
        self.assertEqual(self._valor_por_etiqueta(ws, "Forma de pago"), "30 dias")
        self.assertEqual(
            self._valor_por_etiqueta(ws, "Condicion de la cotizacion"),
            "\n".join(condiciones),
        )
        self.assertEqual(
            self._valor_por_etiqueta(ws, "Observacion al cliente"),
            "Solicitud de presupuesto n 4587 - Lic municipal 221304",
        )

    def test_cotizacion_acepta_forma_pago_personalizada_y_anticipo(self):
        cotizacion_id = self._crear_cotizacion(forma_pago="Anticipo 35%")
        cotizacion = self._cotizacion(cotizacion_id)
        self.assertEqual(cotizacion.forma_pago, "Anticipo 35%")

        html = self.client.get(f"/cotizacion/{cotizacion_id}").get_data(as_text=True)
        self.assertIn("Forma de pago", html)
        self.assertIn("Anticipo 35%", html)

        cotizacion_personalizada_id = self._crear_cotizacion(
            forma_pago="50% al inicio y saldo contra entrega"
        )
        cotizacion_personalizada = self._cotizacion(cotizacion_personalizada_id)
        self.assertEqual(cotizacion_personalizada.forma_pago, "50% al inicio y saldo contra entrega")

    def test_formulario_cotizador_tiene_pago_personalizado_y_anticipo(self):
        response = self.client.get("/")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('id="forma_pago_tipo"', html)
        self.assertIn('value="Personalizado"', html)
        self.assertIn('value="Anticipo"', html)
        self.assertIn('id="forma_pago_personalizada"', html)
        self.assertIn('id="forma_pago_anticipo_pct"', html)
        self.assertIn('name="forma_pago"', html)
        self.assertIn('type="text" name="costo[]"', html)

    def test_condicion_cotizacion_legado_sigue_leyendose_como_unica_condicion(self):
        cotizacion_id = self._crear_cotizacion(
            condicion_cotizacion="Las imagenes son ilustrativas. Oferta sujeta a stock.",
        )

        cotizacion = self._cotizacion(cotizacion_id)
        self.assertEqual(
            cotizacion.condiciones_cotizacion_lista,
            ["Las imagenes son ilustrativas. Oferta sujeta a stock."],
        )

        html = self.client.get(f"/cotizacion/{cotizacion_id}").get_data(as_text=True)
        self.assertIn("Las imagenes son ilustrativas. Oferta sujeta a stock.", html)

    def test_pdf_toma_nombre_visible_del_usuario_logueado(self):
        cotizacion_id = self._crear_cotizacion()
        self._login("operador", "operador123")

        response = self.client.get(f"/cotizacion/{cotizacion_id}")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Lucas Santacruz", html)
        self.assertNotIn("Administrador Principal", html)

    def test_vista_llave_en_mano_muestra_propuesta_memoria_y_condiciones(self):
        condiciones = [
            "Entrega y puesta en marcha a coordinar.",
            "La oferta contempla materiales y mano de obra.",
        ]
        cotizacion_id = self._crear_cotizacion(
            cliente_contacto="Sra. Martin Moreno",
            forma_pago="30 dias",
            condicion_cotizacion=json.dumps(condiciones, ensure_ascii=False),
            observacion_cliente="Se considera instalacion completa y pruebas en sitio.",
            **{
                "row_id[]": ["row-1", "row-2"],
                "item_id[]": ["", ""],
                "imagen_actual[]": ["", ""],
                "desc[]": ["Camara bullet 4MP", "NVR 16 canales"],
                "detalle[]": ["Incluye soporte y conexion.", "Grabacion central y configuracion inicial."],
                "cant[]": ["4", "1"],
                "costo[]": ["70", "220"],
                "iva_compra[]": ["21", "21"],
                "extra[]": ["5", "5"],
                "margen[]": ["60", "60"],
                "descuento[]": ["0", "0"],
                "carga_fiscal[]": ["0", "0"],
                "iva_item[]": ["21", "21"],
            },
        )

        response = self.client.get(f"/cotizacion/{cotizacion_id}/llave-en-mano")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Propuesta Comercial", html)
        self.assertIn("Memoria Descriptiva", html)
        self.assertIn("Propuesta Económica.", html)
        self.assertIn("Sra. Martin Moreno", html)
        self.assertIn("Forma de pago: 30 dias.", html)
        self.assertIn("Se considera instalacion completa y pruebas en sitio.", html)
        self.assertIn("Camara bullet 4MP", html)
        self.assertIn("NVR 16 canales", html)
        self.assertIn("Los precios expresados incluyen IVA.", html)
        self.assertIn("Nota: La factura se pesificara considerando el TC BNA Ventas Divisa", html)
        self.assertNotIn("Familia:", html)
        self.assertNotIn("usuario@example.com", html)
        self.assertIn("Consolidada", self.client.get("/historial").get_data(as_text=True))

    def test_responsable_inscrito_ve_subtotal_neto_e_iva(self):
        cotizacion_id = self._crear_cotizacion(
            cliente_id=str(self.cliente_ri_id),
            cliente="Cliente Responsable",
            cliente_razon_social="Cliente Responsable SA",
            cliente_cuit="30-87654321-0",
            condicion_iva="Responsable Inscrito",
        )

        response = self.client.get(f"/cotizacion/{cotizacion_id}")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Subtotal Neto:", html)
        self.assertIn("IVA 21%:", html)
        self.assertIn("IVA Total:", html)
        self.assertNotIn("Los precios unitarios incluyen IVA", html)

    def test_responsable_inscrito_detalla_iva_por_alicuota_al_final(self):
        cotizacion_id = self._crear_cotizacion(
            cliente_id=str(self.cliente_ri_id),
            cliente="Cliente Responsable",
            cliente_razon_social="Cliente Responsable SA",
            cliente_cuit="30-87654321-0",
            condicion_iva="Responsable Inscrito",
            **{
                "row_id[]": ["row-1", "row-2"],
                "item_id[]": ["", ""],
                "imagen_actual[]": ["", ""],
                "desc[]": ["Equipo principal", "Accesorio"],
                "detalle[]": ["Con IVA 21", "Con IVA 10.5"],
                "cant[]": ["1", "2"],
                "costo[]": ["100", "50"],
                "iva_compra[]": ["21", "10.5"],
                "extra[]": ["0", "0"],
                "margen[]": ["50", "20"],
                "descuento[]": ["0", "0"],
                "carga_fiscal[]": ["0", "0"],
                "iva_item[]": ["21", "10.5"],
            },
        )

        response = self.client.get(f"/cotizacion/{cotizacion_id}")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("IVA 21%:", html)
        self.assertIn("IVA 10.5%:", html)
        self.assertIn("IVA Total:", html)

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            wb = load_workbook(BytesIO(generar_excel_cotizacion(cotizacion)), data_only=True)

        ws = wb.active
        self.assertIsNotNone(self._valor_por_etiqueta(ws, "IVA 21%"))
        self.assertIsNotNone(self._valor_por_etiqueta(ws, "IVA 10.5%"))

    def test_descuento_solo_aparece_en_pdf_si_alguna_fila_lo_usa(self):
        sin_descuento_id = self._crear_cotizacion()
        html_sin_descuento = self.client.get(f"/cotizacion/{sin_descuento_id}").get_data(as_text=True)
        self.assertNotIn("Desc. %", html_sin_descuento)

        con_descuento_id = self._crear_cotizacion(**{"descuento[]": ["10"]})
        html_con_descuento = self.client.get(f"/cotizacion/{con_descuento_id}").get_data(as_text=True)
        cotizacion_con_descuento = self._cotizacion(con_descuento_id)

        self.assertIn("Desc. %", html_con_descuento)
        self.assertAlmostEqual(cotizacion_con_descuento.total_neto, 143.64, places=2)
        self.assertAlmostEqual(cotizacion_con_descuento.total_iva, 30.16, places=2)
        self.assertAlmostEqual(cotizacion_con_descuento.total_final, 173.80, places=2)

    def test_bonificacion_global_descuenta_total_sin_tocar_metricas_por_item(self):
        cotizacion_id = self._crear_cotizacion(**{"bonificacion_cierre_monto": "50"})

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            item = cotizacion.items[0]

            self.assertAlmostEqual(item.precio_venta, 159.60, places=2)
            self.assertAlmostEqual(cotizacion.total_neto, 159.60, places=2)
            self.assertAlmostEqual(cotizacion.total_iva, 33.52, places=2)
            self.assertAlmostEqual(cotizacion.total_carga_fiscal, 7.98, places=2)
            self.assertAlmostEqual(cotizacion.bonificacion_cierre_monto, 50.0, places=2)
            self.assertAlmostEqual(cotizacion.total_final, 143.12, places=2)

            wb = load_workbook(BytesIO(generar_excel_cotizacion(cotizacion)), data_only=True)

        ws = wb.active
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Bonificacion de cierre"), 50.0, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Ganancia neta (Bolsillo)"), 1.87, places=2)
        self.assertAlmostEqual(self._valor_por_etiqueta(ws, "Total a cobrar"), 143.12, places=2)

        html = self.client.get(f"/cotizacion/{cotizacion_id}").get_data(as_text=True)
        self.assertIn("Bonificacion de cierre:", html)
        self.assertIn("- $ 50.00", html)

    def test_recordatorio_actualiza_fechas_de_seguimiento(self):
        cotizacion_id = self._crear_cotizacion()

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            cotizacion.seguimiento_activo = True
            cotizacion.seguimiento_email = "seguimiento@cuencotech.com"
            cotizacion.seguimiento_cada_dias = 7
            cotizacion.seguimiento_proximo_envio = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(days=1)
            db.session.commit()

        with patch.object(cotizador_app, "enviar_mail_recordatorio", return_value=True):
            with app.app_context():
                cotizador_app.procesar_recordatorios_vencidos()

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            self.assertIsNotNone(cotizacion.seguimiento_ultimo_envio)
            self.assertGreater(cotizacion.seguimiento_proximo_envio, cotizacion.seguimiento_ultimo_envio)

    def test_recordatorio_no_se_envia_dos_veces_si_otro_worker_lo_tomo(self):
        cotizacion_id = self._crear_cotizacion()

        with app.app_context():
            cotizacion = db.session.get(Cotizacion, cotizacion_id)
            cotizacion.seguimiento_activo = True
            cotizacion.seguimiento_email = "seguimiento@cuencotech.com"
            cotizacion.seguimiento_cada_dias = 7
            cotizacion.seguimiento_proximo_envio = cotizador_app.datetime.utcnow() - cotizador_app.timedelta(days=1)
            db.session.commit()

        with patch.object(cotizador_app, "enviar_mail_recordatorio", return_value=True) as enviar_mock:
            with app.app_context():
                cotizador_app.procesar_recordatorios_vencidos()
                cotizador_app.procesar_recordatorios_vencidos()

        self.assertEqual(enviar_mock.call_count, 1)

    def test_caso_2_tipo_cambio_guardado_no_se_pisa_con_bna_actual(self):
        cotizacion_id = self._crear_cotizacion()

        payload_bna_nuevo = {
            "ok": True,
            "rate": 1500.0,
            "date": "24/04/2026",
            "time": "12:00",
            "label": "BNA billete vendedor",
        }
        with patch.object(cotizador_app, "obtener_tipo_cambio_oficial_bna", return_value=payload_bna_nuevo):
            response_get = self.client.get(f"/cotizacion/{cotizacion_id}/editar")
            html = response_get.get_data(as_text=True)
            self.assertEqual(response_get.status_code, 200)
            self.assertIn("Tipo guardado en esta cotizacion: 1450.0000", html)
            self.assertIn("Usa actualizar para consultar el BNA actual.", html)

            data = self._form_cotizacion()
            data.pop("tipo_cambio")
            response_post = self.client.post(f"/cotizacion/{cotizacion_id}/editar", data=data, follow_redirects=False)

        self.assertEqual(response_post.status_code, 302)
        cotizacion = self._cotizacion(cotizacion_id)
        self.assertAlmostEqual(cotizacion.tipo_cambio_usado, 1450.0, places=4)

    def test_cantidades_se_guardan_enteras_y_no_permiten_fracciones(self):
        cotizacion_id = self._crear_cotizacion(**{"cant[]": ["2.9"]})
        cotizacion = self._cotizacion(cotizacion_id)

        self.assertEqual(cotizacion.items[0].cantidad, 2)
        self.assertAlmostEqual(cotizacion.total_neto, 319.20, places=2)
        self.assertAlmostEqual(cotizacion.total_final, 386.23, places=2)


if __name__ == "__main__":
    unittest.main()
