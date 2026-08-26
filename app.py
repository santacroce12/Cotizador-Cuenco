import os
import json
import math
import re
import smtplib
import threading
import time
from datetime import datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from html import unescape
from io import BytesIO
from pathlib import Path

import jwt
import requests
import urllib3
from flask import (
    Flask,
    Response,
    flash,
    g,
    has_app_context,
    has_request_context,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageOps, UnidentifiedImageError
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import case, event, func, literal, or_, union_all
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import contains_eager
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def resolver_ruta_configurada(valor, default_path):
    valor = (valor or "").strip()
    if not valor:
        return default_path
    ruta = Path(valor)
    if not ruta.is_absolute():
        ruta = (Path(app.root_path) / ruta).resolve()
    return ruta


DATA_DIR = resolver_ruta_configurada(os.getenv("DATA_DIR"), Path(app.root_path))
DATA_DIR.mkdir(parents=True, exist_ok=True)
db_path = resolver_ruta_configurada(os.getenv("DATABASE_PATH"), DATA_DIR / "database.db")
db_path.parent.mkdir(parents=True, exist_ok=True)
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path.as_posix()}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024
db = SQLAlchemy(app)

NOMBRE_FANTASIA = "Cuenco Tech"
RAZON_SOCIAL = "Cuenco Tech S.A."
CUIT = "30-71831614-2"
DOMICILIO = "Rafael Cubillos 2056, M5500 Godoy Cruz, Mendoza"
DEFAULT_FOLLOWUP_EMAIL = "equipo@example.com"
DEFAULT_SMTP_HOST = "smtp.example.com"
DEFAULT_SMTP_PORT = 465
DEFAULT_SMTP_USERNAME = "usuario@example.com"
DEFAULT_SMTP_FROM = "usuario@example.com"
DEFAULT_APP_BASE_URL = "http://localhost:9000"
DASHBOARD_FIXED_CURRENCY = "USD"
LOCAL_SETTINGS_PATH = resolver_ruta_configurada(os.getenv("LOCAL_SETTINGS_PATH"), Path(app.root_path) / "local_settings.json")


def cargar_local_settings():
    if not LOCAL_SETTINGS_PATH.exists():
        return {}
    try:
        data = json.loads(LOCAL_SETTINGS_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


LOCAL_SETTINGS = cargar_local_settings()
app.config["SECRET_KEY"] = (
    os.getenv("APP_SECRET_KEY")
    or os.getenv("SECRET_KEY")
    or str(LOCAL_SETTINGS.get("APP_SECRET_KEY") or LOCAL_SETTINGS.get("SECRET_KEY") or "").strip()
    or "change-this-secret-in-local-settings"
)
AUTH_COOKIE_SECURE = str(
    os.getenv("AUTH_COOKIE_SECURE") or LOCAL_SETTINGS.get("AUTH_COOKIE_SECURE") or ""
).strip().lower() in ("1", "true", "yes")
QUOTE_SIGNER_NAME = str(
    os.getenv("QUOTE_SIGNER_NAME") or LOCAL_SETTINGS.get("QUOTE_SIGNER_NAME") or "Equipo Comercial Cuenco Tech"
).strip()
QUOTE_SIGNER_ROLE = str(
    os.getenv("QUOTE_SIGNER_ROLE") or LOCAL_SETTINGS.get("QUOTE_SIGNER_ROLE") or "Area Comercial"
).strip()
QUOTE_CONTACT_EMAIL = str(
    os.getenv("QUOTE_CONTACT_EMAIL") or LOCAL_SETTINGS.get("QUOTE_CONTACT_EMAIL") or DEFAULT_SMTP_FROM
).strip()
QUOTE_CONTACT_PHONE = str(
    os.getenv("QUOTE_CONTACT_PHONE") or LOCAL_SETTINGS.get("QUOTE_CONTACT_PHONE") or ""
).strip()
QUOTE_SIGNATURE_IMAGE = str(
    os.getenv("QUOTE_SIGNATURE_IMAGE") or LOCAL_SETTINGS.get("QUOTE_SIGNATURE_IMAGE") or ""
).strip()
QUOTE_FOOTER_NOTE = str(
    os.getenv("QUOTE_FOOTER_NOTE")
    or LOCAL_SETTINGS.get("QUOTE_FOOTER_NOTE")
    or "Precios sujetos a disponibilidad y confirmacion comercial al momento de la aceptacion."
).strip()
QUOTE_USD_FACTURACION_NOTE = (
    "Nota: La factura se pesificara considerando el TC BNA Ventas Divisa del dia habil anterior a la "
    "fecha de emision de la misma. En caso de existir variacion del tipo de cambio al momento de la "
    "real acreditacion del pago con respecto al TC de facturacion, se emitira ND/NC sobre la diferencia "
    "segun corresponda. Por favor corroborar que la OC contenga esta consideracion para evitar el "
    "aplazamiento de inicio de obra por dilaciones administrativas."
)
try:
    QUOTE_VALIDITY_DAYS = int(
        os.getenv("QUOTE_VALIDITY_DAYS") or LOCAL_SETTINGS.get("QUOTE_VALIDITY_DAYS") or 7
    )
except (TypeError, ValueError):
    QUOTE_VALIDITY_DAYS = 7

PLACEHOLDER_PRODUCTO = "placeholder_product.png"
UPLOADS_PRODUCTOS_DIR = resolver_ruta_configurada(
    os.getenv("UPLOADS_PRODUCTOS_DIR"), Path(app.static_folder) / "uploads" / "productos"
)
UPLOADS_PRODUCTOS_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
MAX_IMAGE_UPLOAD_BYTES = 6 * 1024 * 1024
PRODUCT_IMAGE_MAX_DIMENSION = 1400
PRODUCT_IMAGE_TARGET_BYTES = 450 * 1024
PRODUCT_IMAGE_QUALITY_STEPS = (82, 76, 70, 64, 58)
HISTORIAL_PER_PAGE = 20
DASHBOARD_OPERATIVO_PER_PAGE = 15
AUDITORIA_PER_PAGE = 25
ESTADOS_COTIZACION = ("En progreso", "Aceptada", "Rechazada")
FORMAS_PAGO_COTIZACION = (
    "Contado",
    "15 dias fecha fact.",
    "30 dias",
    "60 dias",
    "A convenir",
)
CONDICIONES_COTIZACION = (
    "Las imagenes son ilustrativas.",
    "Oferta sujeta a stock.",
    "Precios sujetos a variacion sin previo aviso.",
    "Entrega y validez a confirmar.",
)
FAMILIAS_COTIZACION = (
    "SEGURIDAD URBANA",
    "PARKING",
    "TRANSPORTE INTELIGENTE",
    "CONECTIVIDAD SATELITAL",
    "SALAS DE CONTROL",
    "SMART CITIES",
)
SECTORES_CLIENTE = {
    "Publico": ("Municipal", "Provincial", "Nacional", "Otro"),
    "Privado": ("Energía", "Agroindustria", "Hotelería", "Educación", "Turismo", "Retail", "Otro"),
}
REMINDER_POLL_SECONDS = 60
_reminder_worker_lock = threading.Lock()
_reminder_worker_started = False
NUMERO_COTIZACION_PREFIX = "CT"
BNA_PERSONAS_URL = "https://www.bna.com.ar/Personas"
BNA_TC_CACHE_SECONDS = 30 * 60
BNA_REQUEST_TIMEOUT_SECONDS = 3
BNA_USD_SOURCE = str(os.getenv("BNA_USD_SOURCE") or LOCAL_SETTINGS.get("BNA_USD_SOURCE") or "billetes_venta").strip().lower()
_bna_exchange_rate_lock = threading.Lock()
_bna_exchange_rate_cache = {"fetched_at": None, "payload": None}
_bna_refresh_lock = threading.Lock()
_bna_refresh_in_progress = False
_bna_refresh_last_error = None


class Cotizacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero_cotizacion = db.Column(db.String(20), unique=True)
    estado = db.Column(db.String(20), default="En progreso")
    seguimiento_activo = db.Column(db.Boolean, default=False)
    seguimiento_email = db.Column(db.String(150))
    seguimiento_cada_dias = db.Column(db.Integer)
    seguimiento_proximo_envio = db.Column(db.DateTime)
    seguimiento_ultimo_envio = db.Column(db.DateTime)
    nombre_fantasia = db.Column(db.String(100))
    razon_social = db.Column(db.String(100))
    cuit = db.Column(db.String(50))
    cliente = db.Column(db.String(100), nullable=False)
    cliente_id = db.Column(db.Integer, db.ForeignKey("cliente.id"))
    cliente_contacto = db.Column(db.String(100))
    cliente_razon_social = db.Column(db.String(100))
    cliente_cuit = db.Column(db.String(50))
    familia = db.Column(db.String(50))
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    moneda = db.Column(db.String(10), default="USD")
    tipo_cambio_usado = db.Column(db.Float)
    condicion_iva = db.Column(db.String(50))
    condicion_cotizacion = db.Column(db.Text)
    forma_pago = db.Column(db.String(120))
    observacion_cliente = db.Column(db.Text)
    carga_fiscal_pct = db.Column(db.Float, default=0.0)
    carga_fiscal_monto = db.Column(db.Float, default=0.0)
    total_carga_fiscal = db.Column(db.Float, default=0.0)
    bonificacion_cierre_monto = db.Column(db.Float, default=0.0)
    total_neto = db.Column(db.Float)
    total_iva = db.Column(db.Float)
    total_final = db.Column(db.Float)
    items = db.relationship("ItemCotizacion", backref="parent", cascade="all, delete-orphan")

    @property
    def condiciones_cotizacion_lista(self):
        return normalizar_condiciones_cotizacion(self.condicion_cotizacion)

    @property
    def condicion_cotizacion_texto(self):
        return formatear_condiciones_cotizacion(self.condicion_cotizacion)


class ItemCotizacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cotizacion_id = db.Column(db.Integer, db.ForeignKey("cotizacion.id"))
    descripcion = db.Column(db.String(200))
    detalle = db.Column(db.Text)
    cantidad = db.Column(db.Float)
    costo_unitario = db.Column(db.Float)
    iva_compra_pct = db.Column(db.Float, default=0.0)
    costo_extra = db.Column(db.Float, default=5.0)
    margen = db.Column(db.Float)
    descuento_pct = db.Column(db.Float, default=0.0)
    carga_fiscal = db.Column(db.Float, default=0.0)
    iva_item = db.Column(db.Float, default=21.0)
    precio_venta = db.Column(db.Float)
    subtotal = db.Column(db.Float)
    imagen_url = db.Column(db.String(500))

    @property
    def iva_porcentaje(self):
        # Alias para compatibilidad con plantillas previas.
        return self.iva_item or 0.0

    @property
    def precio_venta_unitario(self):
        # Alias para plantillas/exportaciones internas.
        return self.precio_venta or 0.0

    @property
    def precio_venta_total(self):
        return (self.precio_venta or 0.0) * (self.cantidad or 0.0)

    @property
    def descuento_porcentaje(self):
        return self.descuento_pct or 0.0

    @property
    def precio_lista_unitario(self):
        costo = self.costo_unitario or 0.0
        extra = (self.costo_extra or 0.0) / 100.0
        margen = self.margen or 0.0
        return costo * (1 + extra) * (1 + margen)

    @property
    def descuento_total(self):
        return max(0.0, self.precio_lista_unitario - (self.precio_venta or 0.0)) * (self.cantidad or 0.0)


class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    razon_social = db.Column(db.String(100))
    cuit = db.Column(db.String(20), unique=True)
    domicilio = db.Column(db.String(200))
    sector = db.Column(db.String(50))
    subsector = db.Column(db.String(50))
    email = db.Column(db.String(100))
    telefono = db.Column(db.String(50))
    condicion_iva = db.Column(db.String(50), default="Consumidor Final")
    cotizaciones = db.relationship("Cotizacion", backref="cliente_ref", lazy=True)


class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    nombre_completo = db.Column(db.String(120))
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def nombre_para_documentos(self):
        return (self.nombre_completo or "").strip() or self.username


class FamiliaCotizacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)
    activa = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Auditoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey("usuario.id"))
    username = db.Column(db.String(50), nullable=False)
    accion = db.Column(db.String(80), nullable=False)
    tipo_entidad = db.Column(db.String(50), nullable=False)
    entidad_id = db.Column(db.Integer)
    entidad_ref = db.Column(db.String(80))
    detalle = db.Column(db.Text)
    fecha = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    usuario = db.relationship("Usuario", backref="auditorias", lazy=True)


class TipoCambioBnaCache(db.Model):
    __tablename__ = "tipo_cambio_bna_cache"

    id = db.Column(db.Integer, primary_key=True)
    payload_json = db.Column(db.Text, nullable=False)
    actualizado_en = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


class SecuenciaCotizacion(db.Model):
    __tablename__ = "secuencia_cotizacion"

    anio = db.Column(db.Integer, primary_key=True)
    ultimo_numero = db.Column(db.Integer, nullable=False, default=0)


def parsear_numero_bna(valor):
    texto = re.sub(r"[^0-9,.\-]", "", str(valor or "").strip())
    if not texto:
        raise ValueError("Valor vacio")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    return float(texto)


def obtener_configuracion_fuente_bna():
    fuente = BNA_USD_SOURCE
    if fuente == "divisas_venta":
        return {
            "source_code": "divisas_venta",
            "tab_id": "divisas",
            "label": "BNA divisa vendedora",
            "item": "Dolar U.S.A",
        }
    return {
        "source_code": "billetes_venta",
        "tab_id": "billetes",
        "label": "BNA billete vendedor",
        "item": "Dolar U.S.A",
    }


def obtener_modo_ssl_bna():
    modo = str(os.getenv("BNA_SSL_MODE") or LOCAL_SETTINGS.get("BNA_SSL_MODE") or "auto").strip().lower()
    return modo if modo in ("strict", "auto", "insecure") else "auto"


def describir_error_tipo_cambio_bna(exc):
    if isinstance(exc, requests.exceptions.SSLError):
        return "No se pudo validar el certificado SSL del BNA."
    if isinstance(exc, requests.exceptions.Timeout):
        return "La consulta al BNA demoro demasiado."
    if isinstance(exc, requests.exceptions.ConnectionError):
        return "No se pudo conectar con el BNA."
    if isinstance(exc, ValueError):
        return "El formato de respuesta del BNA no se pudo interpretar."
    return "No se pudo consultar el tipo de cambio oficial del BNA."


def descargar_html_bna():
    request_kwargs = {
        "headers": {
            "User-Agent": "Mozilla/5.0 (Cotizador Cuenco Tech)",
            "Accept-Language": "es-AR,es;q=0.9",
        },
        "timeout": BNA_REQUEST_TIMEOUT_SECONDS,
    }
    modo_ssl = obtener_modo_ssl_bna()

    if modo_ssl == "insecure":
        response = requests.get(BNA_PERSONAS_URL, verify=False, **request_kwargs)
        response.raise_for_status()
        return response.text, True

    try:
        response = requests.get(BNA_PERSONAS_URL, verify=True, **request_kwargs)
        response.raise_for_status()
        return response.text, False
    except requests.exceptions.SSLError:
        if modo_ssl != "auto":
            raise
        response = requests.get(BNA_PERSONAS_URL, verify=False, **request_kwargs)
        response.raise_for_status()
        return response.text, True


def extraer_tipo_cambio_bna(html):
    config = obtener_configuracion_fuente_bna()
    contenido = unescape(html or "")
    tabla = re.search(
        rf'<div class="tab-pane[^"]*" id="{config["tab_id"]}".*?<table class="table cotizacion">(.*?)</table>',
        contenido,
        re.IGNORECASE | re.DOTALL,
    )
    if not tabla:
        raise ValueError(f"No se encontro la tabla {config['tab_id']} en BNA")

    bloque_tabla = tabla.group(1)
    fecha_match = re.search(r'<th class="fechaCot">\s*([^<]+)\s*</th>', bloque_tabla, re.IGNORECASE)
    fila_match = re.search(
        rf'<td class="tit">\s*{re.escape(config["item"])}\s*</td>\s*<td>\s*([^<]+)\s*</td>\s*<td>\s*([^<]+)\s*</td>',
        bloque_tabla,
        re.IGNORECASE | re.DOTALL,
    )
    if not fila_match:
        raise ValueError(f"No se encontro la fila {config['item']} en BNA")

    compra_raw, venta_raw = fila_match.groups()
    hora_match = re.search(
        rf'<div class="tab-pane[^"]*" id="{config["tab_id"]}".*?Hora Actualizaci[oó]n:\s*([^<]+)</div>',
        contenido,
        re.IGNORECASE | re.DOTALL,
    )

    return {
        "ok": True,
        "rate": parsear_numero_bna(venta_raw),
        "buy_rate": parsear_numero_bna(compra_raw),
        "date": (fecha_match.group(1).strip() if fecha_match else ""),
        "time": (hora_match.group(1).strip() if hora_match else ""),
        "label": config["label"],
        "source_code": config["source_code"],
        "item": config["item"],
        "source_url": BNA_PERSONAS_URL,
    }


def _payload_tipo_cambio_en_memoria():
    with _bna_exchange_rate_lock:
        cache = _bna_exchange_rate_cache.get("payload")
        return dict(cache) if isinstance(cache, dict) else None


def _actualizar_payload_tipo_cambio_en_memoria(payload, actualizado_en=None):
    if not isinstance(payload, dict):
        return
    with _bna_exchange_rate_lock:
        _bna_exchange_rate_cache["payload"] = dict(payload)
        _bna_exchange_rate_cache["fetched_at"] = (actualizado_en or datetime.utcnow()).timestamp()


def obtener_tipo_cambio_bna_cache():
    payload = _payload_tipo_cambio_en_memoria()
    if payload:
        return payload
    if not has_app_context():
        return None

    cache_persistente = db.session.get(TipoCambioBnaCache, 1)
    if not cache_persistente:
        return None
    try:
        payload = json.loads(cache_persistente.payload_json)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None

    actualizado_en = cache_persistente.actualizado_en or datetime.utcnow()
    payload["fetched_at"] = actualizado_en.isoformat(timespec="seconds") + "Z"
    payload["stale"] = (datetime.utcnow() - actualizado_en).total_seconds() >= BNA_TC_CACHE_SECONDS
    _actualizar_payload_tipo_cambio_en_memoria(payload, actualizado_en)
    return dict(payload)


def guardar_tipo_cambio_bna_cache(payload):
    actualizado_en = datetime.utcnow()
    payload_guardado = dict(payload)
    payload_guardado["fetched_at"] = actualizado_en.isoformat(timespec="seconds") + "Z"
    payload_guardado["stale"] = False
    payload_guardado.pop("error", None)

    if not has_app_context():
        _actualizar_payload_tipo_cambio_en_memoria(payload_guardado, actualizado_en)
        return dict(payload_guardado)

    cache_persistente = db.session.get(TipoCambioBnaCache, 1)
    if not cache_persistente:
        cache_persistente = TipoCambioBnaCache(id=1, payload_json="{}", actualizado_en=actualizado_en)
    cache_persistente.payload_json = json.dumps(payload_guardado, ensure_ascii=False)
    cache_persistente.actualizado_en = actualizado_en
    db.session.add(cache_persistente)
    db.session.commit()
    _actualizar_payload_tipo_cambio_en_memoria(payload_guardado, actualizado_en)
    return dict(payload_guardado)


def obtener_tipo_cambio_oficial_bna(force=False):
    cache = obtener_tipo_cambio_bna_cache()
    # Las respuestas web nunca esperan a BNA si ya conocemos una cotizacion,
    # incluso si vencio. El worker de fondo se encarga de refrescarla.
    if not force and cache:
        return cache

    try:
        html, ssl_insecure = descargar_html_bna()
        payload = extraer_tipo_cambio_bna(html)
        payload["ssl_insecure"] = ssl_insecure
        return guardar_tipo_cambio_bna_cache(payload)
    except (requests.RequestException, TimeoutError, ValueError) as exc:
        mensaje_error = describir_error_tipo_cambio_bna(exc)
        cache = obtener_tipo_cambio_bna_cache()
        if cache:
            payload = dict(cache)
            payload["stale"] = True
            payload["error"] = mensaje_error
            return payload
        config = obtener_configuracion_fuente_bna()
        return {
            "ok": False,
            "rate": None,
            "buy_rate": None,
            "date": "",
            "time": "",
            "label": config["label"],
            "source_code": config["source_code"],
            "item": config["item"],
            "source_url": BNA_PERSONAS_URL,
            "error": mensaje_error,
            "stale": False,
            "ssl_insecure": False,
        }


def _actualizar_tipo_cambio_bna_en_segundo_plano(force=False):
    global _bna_refresh_in_progress, _bna_refresh_last_error
    try:
        with app.app_context():
            payload = obtener_tipo_cambio_oficial_bna(force=True)
            with _bna_refresh_lock:
                _bna_refresh_last_error = payload.get("error") if not payload.get("ok") else None
    except Exception as exc:
        with _bna_refresh_lock:
            _bna_refresh_last_error = str(exc)
    finally:
        with _bna_refresh_lock:
            _bna_refresh_in_progress = False


def solicitar_actualizacion_tipo_cambio_bna(force=False):
    global _bna_refresh_in_progress
    cache = obtener_tipo_cambio_bna_cache()
    if not force and cache and not cache.get("stale"):
        return False
    with _bna_refresh_lock:
        if _bna_refresh_in_progress:
            return False
        _bna_refresh_in_progress = True
    threading.Thread(
        target=_actualizar_tipo_cambio_bna_en_segundo_plano,
        kwargs={"force": force},
        name="cotizador-bna-refresh",
        daemon=True,
    ).start()
    return True


def estado_actualizacion_tipo_cambio_bna():
    with _bna_refresh_lock:
        return _bna_refresh_in_progress, _bna_refresh_last_error


def normalizar_tipo_cambio_valor(valor):
    try:
        tasa = float(valor)
    except (TypeError, ValueError):
        return None
    if tasa <= 0:
        return None
    return round(tasa, 4)


def obtener_condicion_cotizacion_default():
    return CONDICIONES_COTIZACION[0]


def obtener_condiciones_cotizacion_default():
    return list(CONDICIONES_COTIZACION)


def normalizar_condicion_cotizacion_item(valor):
    texto = unescape(str(valor or "")).strip()
    if not texto:
        return ""
    texto = re.sub(r"^[\-\u2022]+\s*", "", texto)
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_condiciones_cotizacion(valor):
    if valor is None:
        return []

    candidatos = []
    if isinstance(valor, (list, tuple, set)):
        candidatos = list(valor)
    elif isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return []
        if texto.startswith("["):
            try:
                data = json.loads(texto)
            except (TypeError, ValueError, json.JSONDecodeError):
                data = None
            if isinstance(data, list):
                candidatos = data
            else:
                candidatos = texto.splitlines() if "\n" in texto else [texto]
        else:
            candidatos = texto.splitlines() if "\n" in texto else [texto]
    else:
        candidatos = [valor]

    condiciones = []
    condiciones_vistas = set()
    for candidato in candidatos:
        texto = normalizar_condicion_cotizacion_item(candidato)
        if not texto:
            continue
        clave = texto.casefold()
        if clave in condiciones_vistas:
            continue
        condiciones_vistas.add(clave)
        condiciones.append(texto)
    return condiciones


def serializar_condiciones_cotizacion(valor):
    return json.dumps(normalizar_condiciones_cotizacion(valor), ensure_ascii=False)


def formatear_condiciones_cotizacion(valor, separador="\n"):
    return separador.join(normalizar_condiciones_cotizacion(valor))


def parsear_decimal(valor, default=0.0):
    if valor is None:
        return default
    texto = str(valor).strip().replace(",", ".")
    if not texto:
        return default
    try:
        numero = float(texto)
    except (TypeError, ValueError):
        return default
    return numero if math.isfinite(numero) else default


def normalizar_iva_venta(valor, default=21.0):
    numero = parsear_decimal(valor, default=default)
    for tasa in (21.0, 10.5, 0.0):
        if abs(numero - tasa) < 0.0001:
            return tasa
    return default


def normalizar_forma_pago(valor):
    valor = re.sub(r"\s+", " ", (valor or "").strip())
    if not valor:
        return "A convenir"
    return valor[:120]


def construir_desglose_iva_cotizacion(cotizacion):
    desglose = {21.0: 0.0, 10.5: 0.0, 0.0: 0.0}
    for item in cotizacion.items:
        tasa = float(item.iva_item or 0.0)
        if tasa not in desglose:
            desglose[tasa] = 0.0
        base_neta = (item.precio_venta or 0.0) * (item.cantidad or 0.0)
        desglose[tasa] += base_neta * (tasa / 100.0)
    return [
        {"rate": tasa, "label": f"IVA {str(tasa).replace('.0', '')}%", "amount": round(monto, 2)}
        for tasa, monto in desglose.items()
        if tasa > 0 and round(monto, 2) > 0
    ]


def normalizar_texto_documento(valor):
    return re.sub(r"\s+", " ", str(valor or "").strip())


def formatear_lista_humana(textos, limite=None):
    items = [normalizar_texto_documento(texto) for texto in textos if normalizar_texto_documento(texto)]
    if not items:
        return ""
    if limite and len(items) > limite:
        resto = len(items) - limite
        items = items[:limite] + [f"{resto} item{'s' if resto != 1 else ''} mas"]
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} y {items[1]}"
    return f"{', '.join(items[:-1])} y {items[-1]}"


def construir_contexto_documento_cotizacion(cotizacion):
    current_user = getattr(g, "current_user", None)
    return {
        "cot": cotizacion,
        "condiciones_cotizacion": cotizacion.condiciones_cotizacion_lista,
        "domicilio_empresa": DOMICILIO,
        "mostrar_descuento": any((item.descuento_pct or 0.0) > 0 for item in cotizacion.items),
        "desglose_iva": construir_desglose_iva_cotizacion(cotizacion),
        "precio_incluye_iva": cotizacion.condicion_iva != "Responsable Inscrito",
        "cliente_sector": cotizacion.cliente_ref.sector if cotizacion.cliente_ref else "",
        "cliente_subsector": cotizacion.cliente_ref.subsector if cotizacion.cliente_ref else "",
        "quote_signature": {
            "name": current_user.nombre_para_documentos if current_user else QUOTE_SIGNER_NAME,
            "role": QUOTE_SIGNER_ROLE,
            "email": QUOTE_CONTACT_EMAIL,
            "phone": QUOTE_CONTACT_PHONE,
            "image": QUOTE_SIGNATURE_IMAGE,
            "validity_days": QUOTE_VALIDITY_DAYS,
            "footer_note": QUOTE_FOOTER_NOTE,
        },
    }


def construir_contexto_llave_en_mano(cotizacion):
    items = list(cotizacion.items)
    cliente_destino = cotizacion.cliente_razon_social or cotizacion.cliente or "Cliente"
    familia = normalizar_texto_documento(cotizacion.familia)
    sector = cotizacion.cliente_ref.sector if cotizacion.cliente_ref else ""
    subsector = cotizacion.cliente_ref.subsector if cotizacion.cliente_ref else ""
    cantidad_total = sum(normalizar_cantidad_entera(item.cantidad, default=1) for item in items)
    componentes = formatear_lista_humana([item.descripcion for item in items], limite=4)
    alcance_base = "La presente propuesta comercial contempla la provision, instalacion, configuracion y puesta en marcha"
    alcance_base += " del proyecto cotizado"
    alcance_base += f" para {cliente_destino}"
    if componentes:
        alcance_base += (
            f", incluyendo {cantidad_total} unidad{'es' if cantidad_total != 1 else ''} distribuidas en "
            f"{len(items)} item{'s' if len(items) != 1 else ''}: {componentes}."
        )
    else:
        alcance_base += "."

    memoria_items = []
    for indice, item in enumerate(items, start=1):
        detalle = normalizar_texto_documento(item.detalle)
        descripcion = normalizar_texto_documento(item.descripcion) or f"Item {indice}"
        detalle_base = detalle or "Sin descripcion adicional cargada."
        memoria_items.append(
            {
                "indice": indice,
                "titulo": descripcion,
                "cantidad": normalizar_cantidad_entera(item.cantidad, default=1),
                "detalle": detalle_base,
            }
        )

    alicuotas = sorted({normalizar_iva_venta(item.iva_item or 0.0) for item in items})
    precio_incluye_iva = cotizacion.condicion_iva != "Responsable Inscrito"
    if precio_incluye_iva:
        iva_resumen = "Incl."
        nota_iva = "Los precios expresados incluyen IVA."
        total_economico = cotizacion.total_final or 0.0
    else:
        iva_resumen = f"{str(alicuotas[0]).replace('.0', '')}%" if len(alicuotas) == 1 else "Mixto"
        nota_iva = "Los precios expresados no incluyen el IVA detallado, el cual debera adicionarse."
        total_economico = cotizacion.total_neto or 0.0

    descripcion_referencia = "Provision de materiales, equipamiento y mano de obra"
    if sector and subsector:
        descripcion_referencia += f" ({sector} / {subsector})"
    descripcion_referencia += "."

    notas_especiales = []
    if cotizacion.observacion_cliente:
        notas_especiales.append(normalizar_texto_documento(cotizacion.observacion_cliente))
    if cotizacion.moneda == "USD" and cotizacion.tipo_cambio_usado:
        notas_especiales.append(
            f"Tipo de cambio de referencia: 1 USD = $ {cotizacion.tipo_cambio_usado:,.4f} ARS."
        )

    condiciones_venta = []
    if cotizacion.forma_pago:
        condiciones_venta.append(f"Forma de pago: {cotizacion.forma_pago}.")
    if QUOTE_VALIDITY_DAYS:
        condiciones_venta.append(f"Mantenimiento de oferta: {QUOTE_VALIDITY_DAYS} dias corridos.")
    condiciones_venta.extend(cotizacion.condiciones_cotizacion_lista)
    if cotizacion.moneda == "USD":
        condiciones_venta.append(QUOTE_USD_FACTURACION_NOTE)
    if QUOTE_FOOTER_NOTE:
        condiciones_venta.append(QUOTE_FOOTER_NOTE)

    contexto = construir_contexto_documento_cotizacion(cotizacion)
    contexto.update(
        {
            "propuesta_comercial_texto": alcance_base,
            "memoria_items": memoria_items,
            "iva_resumen": iva_resumen,
            "nota_iva_llave_mano": nota_iva,
            "total_economico_llave_mano": total_economico,
            "descripcion_economica_llave_mano": descripcion_referencia,
            "notas_especiales_llave_mano": notas_especiales,
            "condiciones_venta_llave_mano": condiciones_venta,
        }
    )
    return contexto


def construir_contexto_tipo_cambio_cotizador(cotizacion=None):
    # Renderizar el cotizador no debe depender de una llamada de red al BNA.
    # Si hay un valor en cache lo mostramos, pero la consulta se hace luego de
    # forma asincronica desde el navegador (o al presionar actualizar).
    payload = obtener_tipo_cambio_bna_cache()
    guardado = normalizar_tipo_cambio_valor(cotizacion.tipo_cambio_usado if cotizacion else None)
    actual = normalizar_tipo_cambio_valor(payload.get("rate") if isinstance(payload, dict) else None)
    inicial = guardado or actual or ""

    if guardado:
        estado_texto = f"Tipo guardado en esta cotizacion: {guardado:.4f}."
        estado_error = False
        if payload and payload.get("ok") and actual:
            estado_texto += f" BNA actual: {actual:.4f}"
            if payload.get("date"):
                estado_texto += f" - {payload['date']}"
            if payload.get("time"):
                estado_texto += f" {payload['time']}"
            if payload.get("stale"):
                estado_texto += " - ultimo dato en cache"
            if abs(actual - guardado) > 0.00005:
                estado_texto += ". Usa actualizar para reemplazarlo."
        else:
            estado_texto += " Usa actualizar para consultar el BNA actual."
        return {
            "payload": payload,
            "guardado": guardado,
            "actual": actual,
            "inicial": inicial,
            "estado_texto": estado_texto,
            "estado_error": estado_error,
        }

    if payload and payload.get("ok") and actual:
        estado_texto = payload.get("label") or "BNA oficial"
        if payload.get("date"):
            estado_texto += f" - {payload['date']}"
        if payload.get("time"):
            estado_texto += f" {payload['time']}"
        if payload.get("stale"):
            estado_texto += " - ultimo dato en cache"
        return {
            "payload": payload,
            "guardado": None,
            "actual": actual,
            "inicial": inicial,
            "estado_texto": estado_texto,
            "estado_error": False,
        }

    return {
        "payload": payload,
        "guardado": None,
        "actual": None,
        "inicial": inicial,
        "estado_texto": "El tipo de cambio se cargara desde BNA al abrir el cotizador.",
        "estado_error": False,
    }

def formatear_numero_cotizacion(anio, secuencia):
    return f"{NUMERO_COTIZACION_PREFIX}-{int(anio):04d}-{int(secuencia):04d}"


def parsear_numero_cotizacion(valor):
    numero = (valor or "").strip().upper()
    match = re.fullmatch(rf"{NUMERO_COTIZACION_PREFIX}-(\d{{4}})-(\d+)", numero)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def generar_token_usuario(usuario):
    token = jwt.encode(
        {
            "user_id": usuario.id,
            "exp": datetime.utcnow() + timedelta(hours=24),
        },
        app.config["SECRET_KEY"],
        algorithm="HS256",
    )
    return token if isinstance(token, str) else token.decode("utf-8")


def obtener_usuario_desde_token(token):
    if not token:
        return None
    try:
        data = jwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
    except jwt.PyJWTError:
        return None
    return db.session.get(Usuario, data.get("user_id"))


def obtener_integracion_api_key():
    return str(
        os.getenv("INTEGRACION_API_KEY")
        or LOCAL_SETTINGS.get("INTEGRACION_API_KEY")
        or ""
    ).strip()


def autenticacion_integracion_valida():
    api_key = obtener_integracion_api_key()
    if not api_key:
        return False

    header_key = (request.headers.get("X-API-Key") or "").strip()
    auth_header = (request.headers.get("Authorization") or "").strip()
    bearer_key = auth_header[7:].strip() if auth_header.lower().startswith("bearer ") else ""

    return bool(header_key and header_key == api_key) or bool(bearer_key and bearer_key == api_key)


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get("x-access-token")
        current_user = obtener_usuario_desde_token(token)
        if not current_user:
            if request.endpoint in {
                "agregar_cliente",
                "actualizar_cliente",
                "agregar_familia_api",
                "actualizar_estado_cotizacion",
                "filtrar_historial",
                "eliminar_cotizacion",
                "clonar_cotizacion",
            }:
                return jsonify({"error": "auth_required"}), 401
            return redirect(url_for("login", next=request.path))
        g.current_user = current_user
        return f(*args, **kwargs)

    return decorated


def integration_token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get("x-access-token")
        current_user = obtener_usuario_desde_token(token)
        if current_user:
            g.current_user = current_user
            return f(*args, **kwargs)

        if autenticacion_integracion_valida():
            g.integration_request = True
            return f(*args, **kwargs)

        return jsonify({"error": "auth_required"}), 401

    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        current_user = getattr(g, "current_user", None)
        if not current_user:
            token = request.cookies.get("x-access-token")
            current_user = obtener_usuario_desde_token(token)
            if current_user:
                g.current_user = current_user
        if not current_user:
            return redirect(url_for("login", next=request.path))
        if not current_user.is_admin:
            if request.endpoint == "eliminar_cotizacion":
                return jsonify({"error": "admin_required"}), 403
            return redirect(url_for("index"))
        return f(*args, **kwargs)

    return decorated


@app.context_processor
def inject_current_user():
    return {
        "current_user": getattr(g, "current_user", None),
    }


with app.app_context():
    try:
        db.create_all()
    except OperationalError as exc:
        # Gunicorn puede bootear varios workers a la vez. En SQLite eso puede
        # disparar un "table already exists" transitorio durante el bootstrap.
        # Si ya se creó en otro worker, seguimos con la fase de migración liviana.
        if "already exists" not in str(exc).lower():
            raise
        db.session.rollback()
    for tabla_obsoleta in ("actividad_crm", "oportunidad_crm"):
        db.session.execute(db.text(f"DROP TABLE IF EXISTS {tabla_obsoleta}"))
    db.session.commit()
    columnas_item = [col[1] for col in db.session.execute(db.text("PRAGMA table_info(item_cotizacion)")).fetchall()]
    if "detalle" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN detalle TEXT"))
        db.session.commit()
    if "iva_item" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN iva_item FLOAT DEFAULT 21.0"))
        db.session.commit()
    if "costo_extra" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN costo_extra FLOAT DEFAULT 5.0"))
        db.session.commit()
    if "iva_compra_pct" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN iva_compra_pct FLOAT DEFAULT 0.0"))
        db.session.commit()
    if "descuento_pct" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN descuento_pct FLOAT DEFAULT 0.0"))
        db.session.commit()
    if "carga_fiscal" not in columnas_item:
        db.session.execute(db.text("ALTER TABLE item_cotizacion ADD COLUMN carga_fiscal FLOAT DEFAULT 0.0"))
        db.session.commit()
    columnas_cot = [col[1] for col in db.session.execute(db.text("PRAGMA table_info(cotizacion)")).fetchall()]
    if "moneda" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN moneda VARCHAR(10) DEFAULT 'ARS'"))
        db.session.commit()
    if "tipo_cambio_usado" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN tipo_cambio_usado FLOAT"))
        db.session.commit()
    if "condicion_iva" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN condicion_iva VARCHAR(50)"))
        db.session.commit()
    if "condicion_cotizacion" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN condicion_cotizacion TEXT"))
        db.session.commit()
    if "forma_pago" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN forma_pago VARCHAR(50)"))
        db.session.commit()
    if "observacion_cliente" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN observacion_cliente TEXT"))
        db.session.commit()
    if "numero_cotizacion" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN numero_cotizacion VARCHAR(20)"))
        db.session.commit()
    if "estado" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN estado VARCHAR(20) DEFAULT 'En progreso'"))
        db.session.commit()
    if "seguimiento_activo" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN seguimiento_activo BOOLEAN DEFAULT 0"))
        db.session.commit()
    if "seguimiento_email" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN seguimiento_email VARCHAR(150)"))
        db.session.commit()
    if "seguimiento_cada_dias" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN seguimiento_cada_dias INTEGER"))
        db.session.commit()
    if "seguimiento_proximo_envio" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN seguimiento_proximo_envio DATETIME"))
        db.session.commit()
    if "seguimiento_ultimo_envio" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN seguimiento_ultimo_envio DATETIME"))
        db.session.commit()
    if "cliente_razon_social" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN cliente_razon_social VARCHAR(100)"))
        db.session.commit()
    if "cliente_contacto" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN cliente_contacto VARCHAR(100)"))
        db.session.commit()
    if "cliente_cuit" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN cliente_cuit VARCHAR(50)"))
        db.session.commit()
    if "cliente_id" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN cliente_id INTEGER"))
        db.session.commit()
    if "familia" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN familia VARCHAR(50)"))
        db.session.commit()
    if "carga_fiscal_pct" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN carga_fiscal_pct FLOAT DEFAULT 0.0"))
        db.session.commit()
    if "carga_fiscal_monto" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN carga_fiscal_monto FLOAT DEFAULT 0.0"))
        db.session.commit()
    if "total_carga_fiscal" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN total_carga_fiscal FLOAT DEFAULT 0.0"))
        db.session.commit()
    if "bonificacion_cierre_monto" not in columnas_cot:
        db.session.execute(db.text("ALTER TABLE cotizacion ADD COLUMN bonificacion_cierre_monto FLOAT DEFAULT 0.0"))
        db.session.commit()
    db.session.execute(
        db.text("UPDATE cotizacion SET estado = 'En progreso' WHERE estado IS NULL OR TRIM(estado) = ''")
    )
    db.session.commit()
    columnas_cliente = [col[1] for col in db.session.execute(db.text("PRAGMA table_info(cliente)")).fetchall()]
    if "domicilio" not in columnas_cliente:
        db.session.execute(db.text("ALTER TABLE cliente ADD COLUMN domicilio VARCHAR(200)"))
        db.session.commit()
    if "sector" not in columnas_cliente:
        db.session.execute(db.text("ALTER TABLE cliente ADD COLUMN sector VARCHAR(50)"))
        db.session.commit()
    if "subsector" not in columnas_cliente:
        db.session.execute(db.text("ALTER TABLE cliente ADD COLUMN subsector VARCHAR(50)"))
        db.session.commit()
    if "email" not in columnas_cliente:
        db.session.execute(db.text("ALTER TABLE cliente ADD COLUMN email VARCHAR(100)"))
        db.session.commit()
    if "telefono" not in columnas_cliente:
        db.session.execute(db.text("ALTER TABLE cliente ADD COLUMN telefono VARCHAR(50)"))
        db.session.commit()
    if "condicion_iva" not in columnas_cliente:
        db.session.execute(
            db.text("ALTER TABLE cliente ADD COLUMN condicion_iva VARCHAR(50) DEFAULT 'Consumidor Final'")
        )
        db.session.commit()
    columnas_usuario = [col[1] for col in db.session.execute(db.text("PRAGMA table_info(usuario)")).fetchall()]
    if "nombre_completo" not in columnas_usuario:
        db.session.execute(db.text("ALTER TABLE usuario ADD COLUMN nombre_completo VARCHAR(120)"))
        db.session.commit()
    if "is_admin" not in columnas_usuario:
        db.session.execute(db.text("ALTER TABLE usuario ADD COLUMN is_admin BOOLEAN DEFAULT 0"))
        db.session.commit()
    columnas_familia = [
        col[1] for col in db.session.execute(db.text("PRAGMA table_info(familia_cotizacion)")).fetchall()
    ]
    if "activa" not in columnas_familia:
        db.session.execute(db.text("ALTER TABLE familia_cotizacion ADD COLUMN activa BOOLEAN DEFAULT 1"))
        db.session.commit()
    if "fecha_creacion" not in columnas_familia:
        db.session.execute(db.text("ALTER TABLE familia_cotizacion ADD COLUMN fecha_creacion DATETIME"))
        db.session.commit()
    db.session.execute(db.text("UPDATE familia_cotizacion SET activa = 1 WHERE activa IS NULL"))
    for familia_default in FAMILIAS_COTIZACION:
        db.session.execute(
            db.text("INSERT OR IGNORE INTO familia_cotizacion (nombre, activa, fecha_creacion) VALUES (:nombre, 1, :fecha)"),
            {"nombre": familia_default, "fecha": datetime.utcnow()},
        )
    db.session.commit()
    db.session.execute(db.text("UPDATE usuario SET is_admin = 0 WHERE is_admin IS NULL"))
    db.session.commit()
    if not db.session.query(Usuario.id).filter(Usuario.is_admin.is_(True)).first():
        primer_usuario = Usuario.query.order_by(Usuario.id.asc()).first()
        if primer_usuario:
            primer_usuario.is_admin = True
            db.session.commit()
    cotizaciones_ordenadas = Cotizacion.query.order_by(Cotizacion.fecha.asc(), Cotizacion.id.asc()).all()
    secuencias_por_anio = {}
    hubo_cambios_numeracion = False
    for cot in cotizaciones_ordenadas:
        anio = (cot.fecha or datetime.utcnow()).year
        secuencias_por_anio[anio] = secuencias_por_anio.get(anio, 0) + 1
        numero_esperado = formatear_numero_cotizacion(anio, secuencias_por_anio[anio])
        if (cot.numero_cotizacion or "").strip() != numero_esperado:
            cot.numero_cotizacion = numero_esperado
            hubo_cambios_numeracion = True
    if hubo_cambios_numeracion:
        db.session.commit()
    for anio, ultimo_numero in secuencias_por_anio.items():
        db.session.execute(
            db.text(
                """
                INSERT INTO secuencia_cotizacion (anio, ultimo_numero)
                VALUES (:anio, :ultimo_numero)
                ON CONFLICT(anio) DO UPDATE SET
                    ultimo_numero = MAX(ultimo_numero, excluded.ultimo_numero)
                """
            ),
            {"anio": anio, "ultimo_numero": ultimo_numero},
        )
    db.session.commit()
    for sql in (
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_cotizacion_numero ON cotizacion (numero_cotizacion)",
        "CREATE INDEX IF NOT EXISTS ix_cotizacion_fecha_id ON cotizacion (fecha DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_cotizacion_estado_fecha ON cotizacion (estado, fecha DESC)",
        "CREATE INDEX IF NOT EXISTS ix_cotizacion_moneda_fecha ON cotizacion (moneda, fecha DESC)",
        "CREATE INDEX IF NOT EXISTS ix_cotizacion_familia_fecha ON cotizacion (familia, fecha DESC)",
        "CREATE INDEX IF NOT EXISTS ix_cotizacion_cliente_fecha ON cotizacion (cliente_id, fecha DESC)",
        "CREATE INDEX IF NOT EXISTS ix_cliente_nombre ON cliente (nombre)",
        "CREATE INDEX IF NOT EXISTS ix_cliente_sector ON cliente (sector)",
        "CREATE INDEX IF NOT EXISTS ix_cliente_subsector ON cliente (subsector)",
        "CREATE INDEX IF NOT EXISTS ix_auditoria_fecha_id ON auditoria (fecha DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_item_cotizacion_cotizacion_id ON item_cotizacion (cotizacion_id)",
    ):
        db.session.execute(db.text(sql))
    db.session.commit()


def archivo_imagen_permitido(filename):
    return Path(filename or "").suffix.lower() in ALLOWED_IMAGE_EXTENSIONS


def optimizar_bytes_imagen(raw_bytes):
    if not raw_bytes:
        raise ValueError("La imagen subida esta vacia.")
    if len(raw_bytes) > MAX_IMAGE_UPLOAD_BYTES:
        raise ValueError("Cada imagen debe pesar como maximo 6 MB antes de subirse.")

    try:
        image = Image.open(BytesIO(raw_bytes))
        image = ImageOps.exif_transpose(image)
        if getattr(image, "is_animated", False):
            image.seek(0)
    except (UnidentifiedImageError, OSError):
        raise ValueError("El archivo seleccionado no es una imagen valida.")

    if image.mode in ("RGBA", "LA", "P"):
        flattened = Image.new("RGB", image.size, (255, 255, 255))
        alpha = image.getchannel("A") if "A" in image.getbands() else None
        flattened.paste(image.convert("RGBA"), mask=alpha)
        image = flattened
    elif image.mode != "RGB":
        image = image.convert("RGB")

    image.thumbnail((PRODUCT_IMAGE_MAX_DIMENSION, PRODUCT_IMAGE_MAX_DIMENSION), Image.Resampling.LANCZOS)

    optimized_bytes = None
    for quality in PRODUCT_IMAGE_QUALITY_STEPS:
        buffer = BytesIO()
        image.save(buffer, format="JPEG", quality=quality, optimize=True, progressive=True)
        optimized_bytes = buffer.getvalue()
        if len(optimized_bytes) <= PRODUCT_IMAGE_TARGET_BYTES:
            break

    return optimized_bytes


def preparar_imagen_producto(file_storage):
    if not file_storage or not file_storage.filename:
        return None
    if not archivo_imagen_permitido(file_storage.filename):
        raise ValueError("Formato de imagen no permitido. Usa PNG, JPG, WEBP o GIF.")
    file_storage.stream.seek(0)
    optimized_bytes = optimizar_bytes_imagen(file_storage.read())
    file_storage.stream.seek(0)
    return optimized_bytes


def guardar_imagen_producto(file_storage, descripcion, row_id, optimized_bytes=None):
    if not file_storage or not file_storage.filename:
        return None
    if optimized_bytes is None:
        optimized_bytes = preparar_imagen_producto(file_storage)
    if not optimized_bytes:
        return None

    base = secure_filename(descripcion) or "producto"
    nombre_archivo = f"{datetime.utcnow():%Y%m%d%H%M%S%f}_{secure_filename(str(row_id))}_{base}.jpg"
    destino = UPLOADS_PRODUCTOS_DIR / nombre_archivo
    destino.write_bytes(optimized_bytes)
    return f"uploads/productos/{nombre_archivo}"


def es_ruta_imagen_local(ruta):
    return bool(ruta) and not str(ruta).startswith(("http://", "https://"))


def eliminar_imagen_local(ruta):
    if not es_ruta_imagen_local(ruta):
        return
    archivo = Path(app.static_folder) / str(ruta)
    if archivo.exists():
        archivo.unlink()


def clonar_imagen_local(ruta):
    if not ruta:
        return None
    if not es_ruta_imagen_local(ruta):
        return ruta

    origen = Path(app.static_folder) / str(ruta)
    if not origen.exists() or not origen.is_file():
        return ruta

    base = secure_filename(origen.stem) or "producto"
    extension = origen.suffix or ".jpg"
    nombre_archivo = f"{datetime.utcnow():%Y%m%d%H%M%S%f}_clone_{base}{extension}"
    destino = UPLOADS_PRODUCTOS_DIR / nombre_archivo
    destino.write_bytes(origen.read_bytes())
    return f"uploads/productos/{nombre_archivo}"


def generar_numero_cotizacion(fecha_referencia=None):
    fecha_base = fecha_referencia or datetime.utcnow()
    secuencia = db.session.execute(
        db.text(
            """
            INSERT INTO secuencia_cotizacion (anio, ultimo_numero)
            VALUES (:anio, 1)
            ON CONFLICT(anio) DO UPDATE SET
                ultimo_numero = ultimo_numero + 1
            RETURNING ultimo_numero
            """
        ),
        {"anio": fecha_base.year},
    ).scalar_one()
    return formatear_numero_cotizacion(fecha_base.year, secuencia)


def obtener_config_smtp():
    config_local = {}
    if LOCAL_SETTINGS_PATH.exists():
        try:
            config_local = json.loads(LOCAL_SETTINGS_PATH.read_text(encoding="utf-8"))
            if not isinstance(config_local, dict):
                config_local = {}
        except Exception as exc:
            print(f"[config] No se pudo leer {LOCAL_SETTINGS_PATH.name}: {exc}")
            config_local = {}

    def cfg(nombre, default=""):
        valor = os.getenv(nombre)
        if valor not in (None, ""):
            return valor
        valor_local = config_local.get(nombre)
        if valor_local not in (None, ""):
            return valor_local
        return default

    port_raw = str(cfg("SMTP_PORT", DEFAULT_SMTP_PORT)).strip()
    try:
        port = int(port_raw)
    except ValueError:
        port = DEFAULT_SMTP_PORT
    use_ssl = str(cfg("SMTP_USE_SSL", "")).strip().lower() in ("1", "true", "yes")
    if not use_ssl and port == 465:
        use_ssl = True
    use_tls = str(cfg("SMTP_USE_TLS", "")).strip().lower() in ("1", "true", "yes")
    if not use_tls and not use_ssl and port == 587:
        use_tls = True
    return {
        "host": str(cfg("SMTP_HOST", DEFAULT_SMTP_HOST)).strip(),
        "port": port,
        "username": str(cfg("SMTP_USERNAME", DEFAULT_SMTP_USERNAME)).strip(),
        "password": str(cfg("SMTP_PASSWORD", "") or ""),
        "from_email": str(cfg("SMTP_FROM", cfg("SMTP_USERNAME", DEFAULT_SMTP_FROM)) or DEFAULT_SMTP_FROM).strip(),
        "use_ssl": use_ssl,
        "use_tls": use_tls,
        "base_url": str(cfg("APP_BASE_URL", DEFAULT_APP_BASE_URL)).strip().rstrip("/"),
        "default_to": str(cfg("FOLLOWUP_DEFAULT_TO_EMAIL", DEFAULT_FOLLOWUP_EMAIL)).strip(),
    }


def smtp_esta_configurado():
    config = obtener_config_smtp()
    return bool(config["host"] and config["from_email"] and config["username"] and config["password"])


def auth_cookie_kwargs():
    return {
        "httponly": True,
        "samesite": "Lax",
        "secure": AUTH_COOKIE_SECURE or request.is_secure,
        "max_age": 60 * 60 * 24,
    }


def buscar_usuario_por_username(username):
    username = (username or "").strip()
    if not username:
        return None
    exacto = Usuario.query.filter(Usuario.username == username).first()
    if exacto:
        return exacto
    return Usuario.query.filter(func.lower(Usuario.username) == username.lower()).first()


def registrar_auditoria(accion, tipo_entidad, entidad_id=None, entidad_ref=None, detalle=None, usuario=None, username=None):
    usuario_actual = usuario or getattr(g, "current_user", None)
    username_final = username or (usuario_actual.username if usuario_actual else "sistema")
    usuario_id = usuario_actual.id if usuario_actual else None

    registro = Auditoria(
        usuario_id=usuario_id,
        username=username_final,
        accion=(accion or "").strip(),
        tipo_entidad=(tipo_entidad or "").strip(),
        entidad_id=entidad_id,
        entidad_ref=(entidad_ref or "").strip() or None,
        detalle=(detalle or "").strip() or None,
    )
    db.session.add(registro)
    return registro


def _antes_de_consulta_sql(_conn, _cursor, _statement, _parameters, context, _executemany):
    context._cotizador_query_started_at = time.perf_counter()


def _despues_de_consulta_sql(_conn, _cursor, _statement, _parameters, context, _executemany):
    inicio = getattr(context, "_cotizador_query_started_at", None)
    if inicio is None or not has_request_context():
        return
    duracion_ms = (time.perf_counter() - inicio) * 1000.0
    g.request_sql_query_count = getattr(g, "request_sql_query_count", 0) + 1
    g.request_sql_duration_ms = getattr(g, "request_sql_duration_ms", 0.0) + duracion_ms


with app.app_context():
    if not event.contains(db.engine, "before_cursor_execute", _antes_de_consulta_sql):
        event.listen(db.engine, "before_cursor_execute", _antes_de_consulta_sql)
    if not event.contains(db.engine, "after_cursor_execute", _despues_de_consulta_sql):
        event.listen(db.engine, "after_cursor_execute", _despues_de_consulta_sql)


@app.before_request
def iniciar_metricas_request():
    g.request_started_at = time.perf_counter()
    g.request_sql_query_count = 0
    g.request_sql_duration_ms = 0.0


@app.after_request
def publicar_metricas_request(response):
    inicio = getattr(g, "request_started_at", None)
    if inicio is None:
        return response
    duracion_ms = (time.perf_counter() - inicio) * 1000.0
    consultas = int(getattr(g, "request_sql_query_count", 0))
    sql_ms = float(getattr(g, "request_sql_duration_ms", 0.0))
    response.headers["X-Response-Time-Ms"] = f"{duracion_ms:.2f}"
    response.headers["X-SQL-Query-Count"] = str(consultas)
    response.headers["Server-Timing"] = f'app;dur={duracion_ms:.2f}, sql;dur={sql_ms:.2f};desc="{consultas} queries"'
    app.logger.info(
        "request_metrics endpoint=%s method=%s status=%s duration_ms=%.2f sql_queries=%d sql_ms=%.2f",
        request.endpoint or "unknown",
        request.method,
        response.status_code,
        duracion_ms,
        consultas,
        sql_ms,
    )
    return response


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_error):
    flash("La carga total de archivos excede el maximo permitido para una sola cotizacion.", "danger")
    destino = request.referrer or url_for("index")
    return redirect(destino)


def normalizar_estado_cotizacion(valor):
    valor = (valor or "").strip()
    if not valor:
        return "En progreso"
    for estado in ESTADOS_COTIZACION:
        if valor.lower() == estado.lower():
            return estado
    return None


def normalizar_nombre_familia(valor):
    valor = re.sub(r"\s+", " ", (valor or "").strip())
    if not valor:
        return None
    return valor.upper()[:50]


def obtener_familias_activas():
    return [
        familia.nombre
        for familia in FamiliaCotizacion.query.filter(FamiliaCotizacion.activa.is_(True))
        .order_by(func.lower(FamiliaCotizacion.nombre).asc())
        .all()
    ]


def obtener_familias_para_selector(valor_actual=None):
    familias = obtener_familias_activas()
    familia_actual = normalizar_familia_cotizacion(valor_actual)
    if familia_actual and familia_actual not in familias:
        familias.append(familia_actual)
    return sorted(familias, key=lambda item: item.lower())


def obtener_familias_para_filtros():
    familias = set(obtener_familias_activas())
    for (familia_usada,) in db.session.query(Cotizacion.familia).filter(
        Cotizacion.familia.isnot(None), func.trim(Cotizacion.familia) != ""
    ).distinct():
        familias.add(familia_usada)
    return sorted(familias, key=lambda item: item.lower())


def buscar_familia_por_nombre(nombre):
    nombre_normalizado = normalizar_nombre_familia(nombre)
    if not nombre_normalizado:
        return None
    return FamiliaCotizacion.query.filter(func.lower(FamiliaCotizacion.nombre) == nombre_normalizado.lower()).first()


def normalizar_familia_cotizacion(valor):
    valor = normalizar_nombre_familia(valor)
    if not valor:
        return None
    familia = buscar_familia_por_nombre(valor)
    if familia:
        return familia.nombre
    familia_usada = (
        db.session.query(Cotizacion.familia)
        .filter(Cotizacion.familia.isnot(None), func.lower(Cotizacion.familia) == valor.lower())
        .first()
    )
    if familia_usada:
        return familia_usada[0]
    return None


def normalizar_subsector_dashboard(valor):
    valor = (valor or "").strip()
    if not valor:
        return None
    for subsectores in SECTORES_CLIENTE.values():
        for subsector in subsectores:
            if valor.lower() == subsector.lower():
                return subsector
    return None


def normalizar_sector_cliente(valor):
    valor = (valor or "").strip()
    for sector in SECTORES_CLIENTE:
        if valor.lower() == sector.lower():
            return sector
    return None


def normalizar_subsector_cliente(sector, valor):
    sector_normalizado = normalizar_sector_cliente(sector)
    if not sector_normalizado:
        return None
    valor = (valor or "").strip()
    for subsector in SECTORES_CLIENTE[sector_normalizado]:
        if valor.lower() == subsector.lower():
            return subsector
    return None


def validar_payload_cliente(data, cliente_actual=None):
    nombre = (data.get("nombre") or "").strip()
    cuit = (data.get("cuit") or "").strip()
    sector = normalizar_sector_cliente(data.get("sector"))
    subsector = normalizar_subsector_cliente(sector, data.get("subsector"))

    if not nombre:
        return None, "Nombre requerido"
    if not sector:
        return None, "Sector requerido"
    if not subsector:
        return None, "Subsector invalido para el sector seleccionado"

    if cuit:
        query = Cliente.query.filter(Cliente.cuit == cuit)
        if cliente_actual:
            query = query.filter(Cliente.id != cliente_actual.id)
        repetido = query.first()
        if repetido:
            return None, "Ya existe un cliente con ese CUIT"

    payload = {
        "nombre": nombre,
        "razon_social": (data.get("razon_social") or "").strip(),
        "cuit": cuit,
        "domicilio": (data.get("domicilio") or "").strip(),
        "sector": sector,
        "subsector": subsector,
        "email": (data.get("email") or "").strip(),
        "telefono": (data.get("telefono") or "").strip(),
        "condicion_iva": (data.get("condicion_iva") or "Consumidor Final").strip() or "Consumidor Final",
    }
    return payload, None


def serializar_cliente(cliente):
    return {
        "id": cliente.id,
        "nombre": cliente.nombre,
        "razon_social": cliente.razon_social,
        "cuit": cliente.cuit,
        "domicilio": cliente.domicilio,
        "sector": cliente.sector,
        "subsector": cliente.subsector,
        "email": cliente.email,
        "telefono": cliente.telefono,
        "condicion_iva": cliente.condicion_iva,
        "cotizaciones_count": len(cliente.cotizaciones or []),
    }


def serializar_item_cotizacion(item):
    return {
        "id": item.id,
        "descripcion": item.descripcion,
        "detalle": item.detalle,
        "cantidad": item.cantidad,
        "precio_venta": item.precio_venta,
        "subtotal": item.subtotal,
        "iva_item": item.iva_item,
    }


def serializar_cotizacion(cotizacion, incluir_items=False):
    payload = {
        "id": cotizacion.id,
        "numero_cotizacion": cotizacion.numero_cotizacion or str(cotizacion.id).zfill(4),
        "estado": normalizar_estado_cotizacion(cotizacion.estado) or "En progreso",
        "cliente": cotizacion.cliente,
        "cliente_id": cotizacion.cliente_id,
        "cliente_nombre": (cotizacion.cliente_ref.nombre if cotizacion.cliente_ref else cotizacion.cliente) or "Sin Cliente",
        "cliente_cuit": cotizacion.cliente_cuit or (cotizacion.cliente_ref.cuit if cotizacion.cliente_ref else None),
        "familia": cotizacion.familia or "",
        "fecha": cotizacion.fecha.isoformat() if cotizacion.fecha else None,
        "moneda": cotizacion.moneda or "ARS",
        "total_neto": cotizacion.total_neto or 0.0,
        "total_iva": cotizacion.total_iva or 0.0,
        "total_final": cotizacion.total_final or 0.0,
        "condicion_iva": cotizacion.condicion_iva,
        "forma_pago": cotizacion.forma_pago,
        "observacion_cliente": cotizacion.observacion_cliente,
    }
    if incluir_items:
        payload["items"] = [serializar_item_cotizacion(item) for item in cotizacion.items]
    return payload


def parsear_entero_positivo(valor, default=None):
    if valor is None:
        valor = ""
    valor = str(valor).strip()
    if not valor:
        return default
    try:
        numero = int(valor)
    except ValueError:
        return default
    return numero if numero > 0 else default


def normalizar_cantidad_entera(valor, default=1):
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return default
    if numero < 1:
        return default
    return int(numero)


def paginar_query(query, page=1, per_page=20):
    page = max(parsear_entero_positivo(page, default=1) or 1, 1)
    per_page = max(parsear_entero_positivo(per_page, default=20) or 20, 1)
    total = query.order_by(None).count()
    pages = max((total + per_page - 1) // per_page, 1)
    if page > pages:
        page = pages
    items = query.limit(per_page).offset((page - 1) * per_page).all()
    start = ((page - 1) * per_page) + 1 if total else 0
    end = min(page * per_page, total) if total else 0
    return {
        "items": items,
        "page": page,
        "per_page": per_page,
        "total": total,
        "pages": pages,
        "has_prev": page > 1,
        "has_next": page < pages,
        "prev_page": page - 1 if page > 1 else None,
        "next_page": page + 1 if page < pages else None,
        "start": start,
        "end": end,
    }


def construir_paginas_visibles(page, pages, radius=2):
    if pages <= 1:
        return [1]
    inicio = max(page - radius, 1)
    fin = min(page + radius, pages)
    paginas = list(range(inicio, fin + 1))
    if 1 not in paginas:
        paginas.insert(0, 1)
    if pages not in paginas:
        paginas.append(pages)
    resultado = []
    anterior = None
    for numero in paginas:
        if anterior and numero - anterior > 1:
            resultado.append(None)
        resultado.append(numero)
        anterior = numero
    return resultado


def normalizar_followup_email(valor, cliente_sel=None):
    config = obtener_config_smtp()
    return config["default_to"] or DEFAULT_FOLLOWUP_EMAIL


def construir_link_cotizacion(cotizacion, editar=False):
    config = obtener_config_smtp()
    sufijo = f"/cotizacion/{cotizacion.id}/editar" if editar else f"/cotizacion/{cotizacion.id}"
    return f"{config['base_url']}{sufijo}"


def preparar_seguimiento_cotizacion(cotizacion, cliente_sel=None):
    activo_anterior = bool(cotizacion.seguimiento_activo)
    email_anterior = (cotizacion.seguimiento_email or "").strip()
    dias_anterior = cotizacion.seguimiento_cada_dias
    activo = request.form.get("seguimiento_activo") == "1"
    email_raw = (request.form.get("seguimiento_email") or "").strip()
    dias = parsear_entero_positivo(request.form.get("seguimiento_cada_dias"), default=None)
    email = normalizar_followup_email(email_raw, cliente_sel=cliente_sel)

    if not activo:
        cotizacion.seguimiento_activo = False
        cotizacion.seguimiento_email = email or None
        cotizacion.seguimiento_cada_dias = dias
        cotizacion.seguimiento_proximo_envio = None
        return

    if not dias:
        dias = 7

    cotizacion.seguimiento_activo = True
    cotizacion.seguimiento_email = email or None
    cotizacion.seguimiento_cada_dias = dias
    if (
        not cotizacion.seguimiento_proximo_envio
        or not activo_anterior
        or email_anterior != (email or "")
        or dias_anterior != dias
    ):
        cotizacion.seguimiento_proximo_envio = datetime.utcnow() + timedelta(days=dias)


def enviar_mail_recordatorio(cotizacion):
    config = obtener_config_smtp()
    destinatario = (cotizacion.seguimiento_email or "").strip()
    if not smtp_esta_configurado() or not destinatario:
        return False

    asunto = f"Seguimiento cotizacion #{cotizacion.numero_cotizacion or cotizacion.id}"
    link_ver = construir_link_cotizacion(cotizacion, editar=False)
    link_editar = construir_link_cotizacion(cotizacion, editar=True)
    cuerpo = f"""Hola,

Hay alguna novedad con esta cotizacion?

Cotizacion: #{cotizacion.numero_cotizacion or cotizacion.id}
Cliente: {cotizacion.cliente}
Estado actual: {cotizacion.estado or 'En progreso'}
Total: {cotizacion.moneda or 'ARS'} {cotizacion.total_final or 0:.2f}

Ver cotizacion:
{link_ver}

Editar cotizacion / cambiar estado:
{link_editar}

Este recordatorio se envio automaticamente desde Cuenco Tech.
"""

    mensaje = EmailMessage()
    mensaje["Subject"] = asunto
    mensaje["From"] = config["from_email"]
    mensaje["To"] = destinatario
    mensaje.set_content(cuerpo)

    smtp_class = smtplib.SMTP_SSL if config["use_ssl"] else smtplib.SMTP
    with smtp_class(config["host"], config["port"], timeout=20) as smtp:
        if config["use_tls"] and not config["use_ssl"]:
            smtp.starttls()
        smtp.login(config["username"], config["password"])
        smtp.send_message(mensaje)
    return True


def procesar_recordatorios_vencidos():
    ahora = datetime.utcnow()
    cotizaciones = Cotizacion.query.filter(
        Cotizacion.seguimiento_activo.is_(True),
        Cotizacion.seguimiento_proximo_envio.isnot(None),
        Cotizacion.seguimiento_proximo_envio <= ahora,
    ).all()

    for cotizacion in cotizaciones:
        if normalizar_estado_cotizacion(cotizacion.estado) != "En progreso":
            continue
        if not cotizacion.seguimiento_email:
            continue
        bloqueo_hasta = ahora + timedelta(minutes=10)
        filas_bloqueadas = Cotizacion.query.filter(
            Cotizacion.id == cotizacion.id,
            Cotizacion.seguimiento_activo.is_(True),
            Cotizacion.seguimiento_proximo_envio.isnot(None),
            Cotizacion.seguimiento_proximo_envio <= ahora,
        ).update(
            {Cotizacion.seguimiento_proximo_envio: bloqueo_hasta},
            synchronize_session=False,
        )
        db.session.commit()
        if filas_bloqueadas != 1:
            continue
        cotizacion = db.session.get(Cotizacion, cotizacion.id)
        try:
            enviado = enviar_mail_recordatorio(cotizacion)
        except Exception as exc:
            print(f"[seguimiento] Error enviando cotizacion {cotizacion.id}: {exc}")
            continue
        if not enviado:
            continue

        dias = cotizacion.seguimiento_cada_dias or 7
        ahora_envio = datetime.utcnow()
        cotizacion.seguimiento_ultimo_envio = ahora_envio
        cotizacion.seguimiento_proximo_envio = ahora_envio + timedelta(days=dias)
        db.session.add(cotizacion)

    if cotizaciones:
        db.session.commit()


def worker_recordatorios():
    while True:
        try:
            with app.app_context():
                procesar_recordatorios_vencidos()
                solicitar_actualizacion_tipo_cambio_bna()
        except Exception as exc:
            print(f"[seguimiento] Worker error: {exc}")
        time.sleep(REMINDER_POLL_SECONDS)


def iniciar_worker_recordatorios():
    global _reminder_worker_started
    if _reminder_worker_started:
        return
    with _reminder_worker_lock:
        if _reminder_worker_started:
            return
        thread = threading.Thread(target=worker_recordatorios, name="cotizador-followup-worker", daemon=True)
        thread.start()
        _reminder_worker_started = True


def construir_items_precargados(cotizacion):
    items_precargados = []
    if not cotizacion:
        return items_precargados

    for item in cotizacion.items:
        if item.imagen_url:
            preview_url = item.imagen_url if str(item.imagen_url).startswith(("http://", "https://")) else url_for(
                "static", filename=item.imagen_url
            )
        else:
            preview_url = url_for("static", filename=PLACEHOLDER_PRODUCTO)

        items_precargados.append(
            {
                "item_id": item.id,
                "descripcion": item.descripcion or "",
                "detalle": item.detalle or "",
                "cantidad": normalizar_cantidad_entera(item.cantidad, default=1),
                "costo_unitario": item.costo_unitario or 0,
                "iva_compra_pct": item.iva_compra_pct or 0,
                "costo_extra": item.costo_extra or 0,
                "margen": round((item.margen or 0) * 100, 2),
                "descuento_pct": item.descuento_pct or 0,
                "carga_fiscal": item.carga_fiscal or 0,
                "iva_item": item.iva_item or 0,
                "imagen_url": item.imagen_url or "",
                "preview_url": preview_url,
            }
        )
    return items_precargados


def renderizar_cotizador(cotizacion=None):
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    tipo_cambio_ctx = construir_contexto_tipo_cambio_cotizador(cotizacion)
    condiciones_iniciales = (
        cotizacion.condiciones_cotizacion_lista
        if cotizacion
        else obtener_condiciones_cotizacion_default()
    )
    return render_template(
        "cotizador.html",
        clientes=clientes,
        cotizacion=cotizacion,
        modo_edicion=bool(cotizacion),
        items_precargados=construir_items_precargados(cotizacion),
        familias_cotizacion=obtener_familias_para_selector(cotizacion.familia if cotizacion else None),
        smtp_configurado=smtp_esta_configurado(),
        followup_default_email=obtener_config_smtp()["default_to"],
        tipo_cambio_oficial=tipo_cambio_ctx["payload"],
        tipo_cambio_inicial=tipo_cambio_ctx["inicial"],
        tipo_cambio_guardado=tipo_cambio_ctx["guardado"],
        tipo_cambio_actual=tipo_cambio_ctx["actual"],
        tipo_cambio_estado_texto=tipo_cambio_ctx["estado_texto"],
        tipo_cambio_estado_error=tipo_cambio_ctx["estado_error"],
        formas_pago_cotizacion=FORMAS_PAGO_COTIZACION,
        condicion_cotizacion_texto_inicial=formatear_condiciones_cotizacion(condiciones_iniciales),
    )


def generar_excel_cotizacion(cotizacion):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizacion"

    dark_fill = PatternFill("solid", fgColor="0F172A")
    mid_fill = PatternFill("solid", fgColor="E2E8F0")
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    title_font = Font(size=15, bold=True)
    thin_side = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_row(row_idx, fill=None, font=None, alignment=None):
        for cell in ws[row_idx]:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            cell.border = border

    def write_label_value(row_idx, label, value):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).font = bold_font
        ws.cell(row=row_idx, column=1).fill = mid_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2).border = border

    ws.merge_cells("A1:N1")
    ws["A1"] = f"Cotizacion {cotizacion.numero_cotizacion or cotizacion.id}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = dark_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].border = border

    row = 3
    datos_generales = [
        ("Empresa", cotizacion.nombre_fantasia or NOMBRE_FANTASIA),
        ("Razon social empresa", cotizacion.razon_social or RAZON_SOCIAL),
        ("CUIT empresa", cotizacion.cuit or CUIT),
        ("Domicilio", DOMICILIO),
        ("Numero de cotizacion", cotizacion.numero_cotizacion or cotizacion.id),
        ("Fecha de creacion", cotizacion.fecha.strftime("%d/%m/%Y %H:%M") if cotizacion.fecha else ""),
        ("Estado", cotizacion.estado or "En progreso"),
        ("Familia", cotizacion.familia or ""),
        ("Cliente", cotizacion.cliente or ""),
        ("Contacto cliente", cotizacion.cliente_contacto or ""),
        ("Razon social cliente", cotizacion.cliente_razon_social or ""),
        ("CUIT cliente", cotizacion.cliente_cuit or ""),
        ("Moneda", cotizacion.moneda or "ARS"),
        ("Tipo de cambio usado", cotizacion.tipo_cambio_usado if cotizacion.tipo_cambio_usado is not None else ""),
        ("Condicion IVA", cotizacion.condicion_iva or ""),
        ("Forma de pago", cotizacion.forma_pago or ""),
        ("Condicion de la cotizacion", formatear_condiciones_cotizacion(cotizacion.condicion_cotizacion)),
        ("Observacion al cliente", cotizacion.observacion_cliente or ""),
    ]
    for label, value in datos_generales:
        write_label_value(row, label, value)
        row += 1

    row += 1
    headers = [
        "Producto / Servicio",
        "Descripcion",
        "Cantidad",
        "Costo neto unitario",
        "IVA compra %",
        "Otros %",
        "Margen %",
        "Descuento %",
        "Carga fiscal %",
        "IVA %",
        "Ganancia neta U.",
        "Precio venta unitario",
        "Precio venta total",
        "Subtotal final",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_row(row, fill=dark_fill, font=white_font, alignment=Alignment(horizontal="center"))

    money_columns = {4, 11, 12, 13, 14}
    percentage_columns = {5, 6, 7, 8, 9, 10}
    costo_neto_real = 0.0
    iva_compra_total = 0.0
    otros_costos_total = 0.0
    for item in cotizacion.items:
        row += 1
        cantidad = normalizar_cantidad_entera(item.cantidad, default=1)
        costo_unitario = item.costo_unitario or 0.0
        iva_compra_pct = item.iva_compra_pct or 0.0
        extra_pct = item.costo_extra or 0.0
        costo_real_unitario = costo_unitario * (1 + (extra_pct / 100.0))
        ganancia_neta_unitaria = (item.precio_venta or 0.0) - costo_real_unitario - ((item.precio_venta or 0.0) * ((item.carga_fiscal or 0.0) / 100.0))
        costo_base_total = cantidad * costo_unitario
        iva_compra_total += costo_base_total * (iva_compra_pct / 100.0)
        otros_costos_total += costo_base_total * (extra_pct / 100.0)
        costo_neto_real += costo_base_total + (costo_base_total * (extra_pct / 100.0))
        values = [
            item.descripcion or "",
            item.detalle or "",
            cantidad,
            costo_unitario,
            iva_compra_pct,
            extra_pct,
            (item.margen or 0) * 100,
            item.descuento_pct or 0,
            item.carga_fiscal or 0,
            item.iva_item or 0,
            ganancia_neta_unitaria,
            item.precio_venta or 0,
            item.precio_venta_total or 0,
            item.subtotal or 0,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col_idx in money_columns:
                cell.number_format = '#,##0.00'
            elif col_idx in percentage_columns:
                cell.number_format = '0.00'

    row += 2
    resumen_inicio = row
    subtotal_neto = cotizacion.total_neto or 0.0
    descuento_total = round(sum((item.descuento_total or 0.0) for item in cotizacion.items), 2)
    total_carga_fiscal = cotizacion.total_carga_fiscal or 0.0
    bonificacion_cierre = cotizacion.bonificacion_cierre_monto or 0.0
    ganancia_neta_bolsillo = subtotal_neto - costo_neto_real - total_carga_fiscal - bonificacion_cierre
    iva_a_pagar = (cotizacion.total_iva or 0.0) - iva_compra_total
    resumen = [
        ("Costo neto real (Base + Otros)", round(costo_neto_real, 2)),
        ("IVA compra credito", round(iva_compra_total, 2)),
        ("IVA a pagar estimado", round(iva_a_pagar, 2)),
        ("Descuento total", descuento_total),
        ("Subtotal venta neto", subtotal_neto),
        ("Bonificacion de cierre", round(bonificacion_cierre, 2)),
        ("Carga fiscal retenida", round(total_carga_fiscal, 2)),
        ("Ganancia neta (Bolsillo)", round(ganancia_neta_bolsillo, 2)),
        ("IVA total", cotizacion.total_iva or 0),
        ("Total a cobrar", cotizacion.total_final or 0),
    ]
    for iva_linea in construir_desglose_iva_cotizacion(cotizacion):
        resumen.insert(-1, (iva_linea["label"], iva_linea["amount"]))
    for idx, (label, value) in enumerate(resumen):
        r = resumen_inicio + idx
        ws.cell(row=r, column=11, value=label)
        ws.cell(row=r, column=12, value=value)
        ws.cell(row=r, column=11).font = bold_font
        ws.cell(row=r, column=11).fill = total_fill
        ws.cell(row=r, column=11).border = border
        ws.cell(row=r, column=12).border = border
        ws.cell(row=r, column=12).number_format = '#,##0.00'
        if label == "Total a cobrar":
            ws.cell(row=r, column=11).font = Font(bold=True, color="0F172A")
            ws.cell(row=r, column=12).font = Font(bold=True, color="0F172A")

    for column, width in {
        "A": 28,
        "B": 34,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 18,
        "K": 18,
        "L": 18,
        "M": 16,
        "N": 16,
    }.items():
        ws.column_dimensions[column].width = width

    for current_row in ws.iter_rows():
        for cell in current_row:
            if cell.row == 1:
                continue
            if not cell.alignment:
                cell.alignment = Alignment(vertical="center")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def aplicar_filtro_fecha_cotizaciones(query, desde="", hasta=""):
    if desde:
        query = query.filter(func.date(Cotizacion.fecha) >= desde)
    if hasta:
        query = query.filter(func.date(Cotizacion.fecha) <= hasta)
    return query


def parsear_fecha_iso(valor):
    valor = (valor or "").strip()
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        return None


def resolver_rango_dashboard(periodo, desde_raw="", hasta_raw=""):
    hoy = datetime.now().date()
    periodo = (periodo or "30").strip().lower()
    rangos_rapidos = {
        "7": 6,
        "30": 29,
        "60": 59,
        "90": 89,
        "365": 364,
    }

    if periodo in rangos_rapidos:
        hasta_date = hoy
        desde_date = hoy - timedelta(days=rangos_rapidos[periodo])
    else:
        periodo = "custom"
        desde_date = parsear_fecha_iso(desde_raw)
        hasta_date = parsear_fecha_iso(hasta_raw) or hoy
        if desde_date is None:
            desde_date = hasta_date - timedelta(days=29)

    if desde_date > hasta_date:
        desde_date, hasta_date = hasta_date, desde_date

    return periodo, desde_date.isoformat(), hasta_date.isoformat(), desde_date, hasta_date


def calcular_variacion_porcentual(actual, anterior):
    if not anterior:
        return None if not actual else 100.0
    return round(((actual - anterior) / anterior) * 100.0, 1)


def construir_resumen_dashboard_sql(query):
    fila = (
        query.order_by(None)
        .with_entities(
            func.count(Cotizacion.id).label("total"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Aceptada", 1), else_=0)),
                0,
            ).label("aceptadas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Rechazada", 1), else_=0)),
                0,
            ).label("rechazadas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "En progreso", 1), else_=0)),
                0,
            ).label("en_progreso"),
            func.coalesce(func.sum(Cotizacion.total_final), 0.0).label("total_importe"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "En progreso", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("importe_pipeline"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Aceptada", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("importe_aceptado"),
        )
        .one()
    )
    return {
        "total": int(fila.total or 0),
        "aceptadas": int(fila.aceptadas or 0),
        "rechazadas": int(fila.rechazadas or 0),
        "en_progreso": int(fila.en_progreso or 0),
        "total_importe": round(float(fila.total_importe or 0.0), 2),
        "importe_pipeline": round(float(fila.importe_pipeline or 0.0), 2),
        "importe_aceptado": round(float(fila.importe_aceptado or 0.0), 2),
    }


def construir_series_dashboard_sql(query, desde_date, hasta_date):
    total_dias = max((hasta_date - desde_date).days + 1, 1)
    usar_semanas = total_dias > 62
    labels = []
    buckets = []

    cursor = desde_date
    while cursor <= hasta_date:
        if usar_semanas:
            bucket_start = cursor
            bucket_end = min(cursor + timedelta(days=6), hasta_date)
            labels.append(f"{bucket_start.strftime('%d/%m')} - {bucket_end.strftime('%d/%m')}")
            buckets.append({"start": bucket_start, "end": bucket_end, "creadas": 0, "cerradas": 0, "aceptadas": 0})
            cursor = bucket_end + timedelta(days=1)
        else:
            labels.append(cursor.strftime("%d/%m"))
            buckets.append({"start": cursor, "end": cursor, "creadas": 0, "cerradas": 0, "aceptadas": 0})
            cursor += timedelta(days=1)

    fecha_sql = func.date(Cotizacion.fecha)
    filas = (
        query.order_by(None)
        .with_entities(
            fecha_sql.label("fecha"),
            func.count(Cotizacion.id).label("creadas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado.in_(("Aceptada", "Rechazada")), 1), else_=0)),
                0,
            ).label("cerradas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Aceptada", 1), else_=0)),
                0,
            ).label("aceptadas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Rechazada", 1), else_=0)),
                0,
            ).label("rechazadas"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "En progreso", 1), else_=0)),
                0,
            ).label("en_progreso"),
            func.coalesce(func.sum(Cotizacion.total_final), 0.0).label("total_importe"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "En progreso", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("importe_pipeline"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Aceptada", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("importe_aceptado"),
        )
        .group_by(fecha_sql)
        .all()
    )

    resumen = {
        "total": 0,
        "aceptadas": 0,
        "rechazadas": 0,
        "en_progreso": 0,
        "total_importe": 0.0,
        "importe_pipeline": 0.0,
        "importe_aceptado": 0.0,
    }
    for fila in filas:
        fecha_cot = fila.fecha
        if isinstance(fecha_cot, str):
            fecha_cot = parsear_fecha_iso(fecha_cot)
        elif isinstance(fecha_cot, datetime):
            fecha_cot = fecha_cot.date()
        if fecha_cot is None:
            continue
        if fecha_cot < desde_date or fecha_cot > hasta_date:
            continue
        if usar_semanas:
            index = (fecha_cot - desde_date).days // 7
        else:
            index = (fecha_cot - desde_date).days
        if index < 0 or index >= len(buckets):
            continue

        bucket = buckets[index]
        bucket["creadas"] += int(fila.creadas or 0)
        bucket["cerradas"] += int(fila.cerradas or 0)
        bucket["aceptadas"] += int(fila.aceptadas or 0)
        resumen["total"] += int(fila.creadas or 0)
        resumen["aceptadas"] += int(fila.aceptadas or 0)
        resumen["rechazadas"] += int(fila.rechazadas or 0)
        resumen["en_progreso"] += int(fila.en_progreso or 0)
        resumen["total_importe"] += float(fila.total_importe or 0.0)
        resumen["importe_pipeline"] += float(fila.importe_pipeline or 0.0)
        resumen["importe_aceptado"] += float(fila.importe_aceptado or 0.0)

    for clave in ("total_importe", "importe_pipeline", "importe_aceptado"):
        resumen[clave] = round(resumen[clave], 2)

    return {
        "labels": labels,
        "creadas": [bucket["creadas"] for bucket in buckets],
        "cerradas": [bucket["cerradas"] for bucket in buckets],
        "aceptadas": [bucket["aceptadas"] for bucket in buckets],
        "granularidad": "Semanal" if usar_semanas else "Diaria",
        "resumen": resumen,
    }


def _consulta_agrupacion_dashboard_sql(query, nombre_sql, grupo=None):
    cantidad_sql = func.count(Cotizacion.id)
    total_sql = func.coalesce(func.sum(Cotizacion.total_final), 0.0)
    columnas = []
    if grupo is not None:
        columnas.append(literal(grupo).label("grupo"))
    columnas.extend(
        (
            nombre_sql.label("nombre"),
            cantidad_sql.label("cantidad"),
            func.coalesce(
                func.sum(case((Cotizacion.estado == "Aceptada", 1), else_=0)),
                0,
            ).label("aceptadas"),
            total_sql.label("total"),
            func.coalesce(
                func.sum(case((Cotizacion.moneda == "ARS", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("total_ars"),
            func.coalesce(
                func.sum(case((Cotizacion.moneda == "USD", Cotizacion.total_final), else_=0.0)),
                0.0,
            ).label("total_usd"),
        )
    )
    return query.order_by(None).with_entities(*columnas).group_by(nombre_sql)


def _agrupar_dashboard_sql(query, nombre_sql, limite=None):
    consulta = _consulta_agrupacion_dashboard_sql(query, nombre_sql)
    cantidad_sql = func.count(Cotizacion.id)
    total_sql = func.coalesce(func.sum(Cotizacion.total_final), 0.0)
    consulta = consulta.order_by(cantidad_sql.desc(), total_sql.desc(), func.lower(nombre_sql).asc())
    if limite:
        consulta = consulta.limit(limite)
    return consulta.all()


def _serializar_agrupacion_dashboard(filas, total_cotizaciones=None):
    resultado = []
    for fila in filas:
        cantidad = int(fila.cantidad or 0)
        item = {
            "nombre": fila.nombre,
            "cantidad": cantidad,
            "aceptadas": int(fila.aceptadas or 0),
            "total": round(float(fila.total or 0.0), 2),
            "total_ars": round(float(fila.total_ars or 0.0), 2),
            "total_usd": round(float(fila.total_usd or 0.0), 2),
        }
        if total_cotizaciones is not None:
            item["porcentaje"] = (
                round((cantidad / total_cotizaciones) * 100.0, 1) if total_cotizaciones else 0.0
            )
        resultado.append(item)
    return resultado


def construir_top_clientes_dashboard_sql(query, limite=5):
    nombre_cliente_sql = case(
        (
            Cliente.id.isnot(None),
            func.coalesce(func.nullif(func.trim(Cliente.nombre), ""), "Sin cliente"),
        ),
        else_=func.coalesce(func.nullif(func.trim(Cotizacion.cliente), ""), "Sin cliente"),
    )
    return _serializar_agrupacion_dashboard(
        _agrupar_dashboard_sql(query, nombre_cliente_sql, limite=limite)
    )


def construir_desglose_dashboard_sql(query, campo_sql, total_cotizaciones, default_label="Sin definir", limite=None):
    nombre_sql = func.coalesce(func.nullif(func.trim(campo_sql), ""), default_label)
    return _serializar_agrupacion_dashboard(
        _agrupar_dashboard_sql(query, nombre_sql, limite=limite),
        total_cotizaciones=total_cotizaciones,
    )


def construir_agrupaciones_dashboard_sql(query, total_cotizaciones):
    nombre_cliente_sql = case(
        (
            Cliente.id.isnot(None),
            func.coalesce(func.nullif(func.trim(Cliente.nombre), ""), "Sin cliente"),
        ),
        else_=func.coalesce(func.nullif(func.trim(Cotizacion.cliente), ""), "Sin cliente"),
    )
    especificaciones = (
        ("clientes", nombre_cliente_sql, None, 5),
        ("familias", func.coalesce(func.nullif(func.trim(Cotizacion.familia), ""), "Sin familia"), total_cotizaciones, 6),
        ("sectores", func.coalesce(func.nullif(func.trim(Cliente.sector), ""), "Sin sector"), total_cotizaciones, None),
        ("subsectores", func.coalesce(func.nullif(func.trim(Cliente.subsector), ""), "Sin subsector"), total_cotizaciones, 8),
    )
    consultas = [
        _consulta_agrupacion_dashboard_sql(query, nombre_sql, grupo=grupo).statement
        for grupo, nombre_sql, _, _ in especificaciones
    ]
    filas = db.session.execute(union_all(*consultas)).all()
    filas_por_grupo = {grupo: [] for grupo, _, _, _ in especificaciones}
    for fila in filas:
        filas_por_grupo[fila.grupo].append(fila)

    resultado = {}
    for grupo, _, total_grupo, limite in especificaciones:
        items = _serializar_agrupacion_dashboard(filas_por_grupo[grupo], total_cotizaciones=total_grupo)
        items.sort(key=lambda item: (-item["cantidad"], -item["total"], item["nombre"].lower()))
        resultado[grupo] = items[:limite] if limite else items
    return resultado


def aplicar_filtro_estado_dashboard(query, estado):
    estado = (estado or "todos").strip().lower()
    if estado == "cerradas":
        estado = "aceptadas"
    if estado == "aceptadas":
        return query.filter(Cotizacion.estado == "Aceptada")
    if estado == "rechazadas":
        return query.filter(Cotizacion.estado == "Rechazada")
    if estado == "en_progreso":
        return query.filter(Cotizacion.estado == "En progreso")
    return query


def aplicar_filtros_segmentacion_dashboard(query, familia="", sector="", subsector="", moneda=""):
    if familia:
        query = query.filter(Cotizacion.familia == familia)
    if sector:
        query = query.filter(Cliente.sector == sector)
    if subsector:
        query = query.filter(Cliente.subsector == subsector)
    if moneda in ("ARS", "USD"):
        query = query.filter(Cotizacion.moneda == moneda)
    return query


def aplicar_filtro_cliente_dashboard(query, cliente=""):
    cliente = (cliente or "").strip()
    if not cliente:
        return query
    patron = f"%{cliente}%"
    return query.filter(
        or_(
            Cotizacion.cliente.ilike(patron),
            Cotizacion.cliente_razon_social.ilike(patron),
            Cotizacion.cliente_cuit.ilike(patron),
            Cliente.nombre.ilike(patron),
            Cliente.razon_social.ilike(patron),
            Cliente.cuit.ilike(patron),
        )
    )


def resolver_estado_dashboard(valor):
    estado = (valor or "todos").strip().lower()
    if estado == "cerradas":
        estado = "aceptadas"
    if estado not in ("todos", "en_progreso", "aceptadas", "rechazadas"):
        estado = "todos"
    return estado


def resolver_filtros_dashboard_request():
    estado = resolver_estado_dashboard(request.args.get("estado"))
    familia = normalizar_familia_cotizacion(request.args.get("familia"))
    op_familia = normalizar_familia_cotizacion(request.args.get("op_familia"))
    sector = normalizar_sector_cliente(request.args.get("sector"))
    subsector = (
        normalizar_subsector_cliente(sector, request.args.get("subsector"))
        if sector
        else normalizar_subsector_dashboard(request.args.get("subsector"))
    )
    moneda = DASHBOARD_FIXED_CURRENCY
    cliente = (request.args.get("cliente") or "").strip()

    periodo, desde, hasta, desde_date, hasta_date = resolver_rango_dashboard(
        request.args.get("periodo"),
        request.args.get("desde"),
        request.args.get("hasta"),
    )

    return {
        "estado": estado,
        "familia": familia,
        "op_familia": op_familia,
        "sector": sector,
        "subsector": subsector,
        "moneda": moneda,
        "cliente": cliente,
        "periodo": periodo,
        "desde": desde,
        "hasta": hasta,
        "desde_date": desde_date,
        "hasta_date": hasta_date,
        "op_page": parsear_entero_positivo(request.args.get("op_page"), default=1) or 1,
    }


def construir_query_dashboard_periodo(filtros):
    query_periodo = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    query_periodo = aplicar_filtro_fecha_cotizaciones(query_periodo, filtros["desde"], filtros["hasta"])
    query_periodo = aplicar_filtro_cliente_dashboard(query_periodo, filtros["cliente"])
    query_periodo = aplicar_filtros_segmentacion_dashboard(
        query_periodo,
        familia=filtros["familia"],
        sector=filtros["sector"],
        subsector=filtros["subsector"],
        moneda=filtros["moneda"],
    )
    return query_periodo


def construir_contexto_dashboard_operativo(query_periodo, filtros):
    query_tabla = query_periodo
    if filtros.get("op_familia"):
        query_tabla = query_tabla.filter(Cotizacion.familia == filtros["op_familia"])
    query_tabla = aplicar_filtro_estado_dashboard(query_tabla, filtros["estado"])
    query_tabla = query_tabla.options(contains_eager(Cotizacion.cliente_ref))
    paginacion = paginar_query(
        query_tabla.order_by(Cotizacion.fecha.desc(), Cotizacion.id.desc()),
        page=filtros.get("op_page", 1),
        per_page=DASHBOARD_OPERATIVO_PER_PAGE,
    )
    return {
        "cotizaciones": paginacion["items"],
        "paginacion_operativa": {
            **paginacion,
            "pages_visible": construir_paginas_visibles(paginacion["page"], paginacion["pages"]),
        },
        "filtro_estado": filtros["estado"],
        "filtro_periodo": filtros["periodo"],
        "filtro_desde": filtros["desde"],
        "filtro_hasta": filtros["hasta"],
        "filtro_familia": filtros["familia"] or "",
        "filtro_op_familia": filtros["op_familia"] or "",
        "filtro_sector": filtros["sector"] or "",
        "filtro_subsector": filtros["subsector"] or "",
        "filtro_cliente": filtros["cliente"] or "",
        "filtro_op_page": paginacion["page"],
    }


def persistir_cotizacion_desde_form(cotizacion=None):
    cliente_id_raw = (request.form.get("cliente_id") or "").strip()
    cliente_sel = None
    if cliente_id_raw.isdigit():
        cliente_sel = db.session.get(Cliente, int(cliente_id_raw))

    moneda = (request.form.get("moneda") or "USD").upper()
    if moneda not in ("ARS", "USD"):
        moneda = "USD"
    tipo_cambio_usado = normalizar_tipo_cambio_valor(request.form.get("tipo_cambio"))
    if tipo_cambio_usado is None:
        tipo_cambio_usado = normalizar_tipo_cambio_valor(cotizacion.tipo_cambio_usado if cotizacion else None)
    if tipo_cambio_usado is None:
        tipo_cambio_usado = normalizar_tipo_cambio_valor(obtener_tipo_cambio_oficial_bna().get("rate"))

    condicion_iva = request.form.get("condicion_iva") or "Consumidor Final"
    if condicion_iva not in ("Exento", "Consumidor Final", "Responsable Inscrito"):
        condicion_iva = "Consumidor Final"
    if cliente_sel and cliente_sel.condicion_iva in ("Exento", "Consumidor Final", "Responsable Inscrito"):
        condicion_iva = cliente_sel.condicion_iva
    forma_pago = normalizar_forma_pago(request.form.get("forma_pago"))
    condiciones_cotizacion_raw = request.form.get("condicion_cotizacion")
    if condiciones_cotizacion_raw is None:
        condiciones_cotizacion = obtener_condiciones_cotizacion_default()
    else:
        condiciones_cotizacion = normalizar_condiciones_cotizacion(condiciones_cotizacion_raw)
    condicion_cotizacion = serializar_condiciones_cotizacion(condiciones_cotizacion) if condiciones_cotizacion else ""
    observacion_cliente = (request.form.get("observacion_cliente") or "").strip()
    cliente_contacto = (request.form.get("cliente_contacto") or "").strip()
    bonificacion_cierre_solicitada = max(0.0, parsear_decimal(request.form.get("bonificacion_cierre_monto") or 0))

    es_edicion = cotizacion is not None
    estado_anterior = None
    familia_anterior = normalizar_familia_cotizacion(cotizacion.familia if cotizacion else None)
    if not cotizacion:
        fecha_creacion = datetime.utcnow()
        cotizacion = Cotizacion(
            fecha=fecha_creacion,
            estado="En progreso",
            nombre_fantasia=NOMBRE_FANTASIA,
            razon_social=RAZON_SOCIAL,
            cuit=CUIT,
            tipo_cambio_usado=tipo_cambio_usado,
            forma_pago=forma_pago,
            condicion_cotizacion=condicion_cotizacion,
            observacion_cliente=observacion_cliente,
            carga_fiscal_pct=0.0,
            carga_fiscal_monto=0.0,
            total_carga_fiscal=0.0,
            bonificacion_cierre_monto=0.0,
            total_neto=0.0,
            total_iva=0.0,
            total_final=0.0,
        )
    else:
        estado_anterior = normalizar_estado_cotizacion(cotizacion.estado) or "En progreso"
        cotizacion.estado = normalizar_estado_cotizacion(cotizacion.estado) or "En progreso"
        estado_form = normalizar_estado_cotizacion(request.form.get("estado"))
        if estado_form:
            cotizacion.estado = estado_form

    cotizacion.nombre_fantasia = NOMBRE_FANTASIA
    cotizacion.razon_social = RAZON_SOCIAL
    cotizacion.cuit = CUIT
    cotizacion.cliente_id = cliente_sel.id if cliente_sel else None
    cotizacion.cliente = cliente_sel.nombre if cliente_sel else (request.form.get("cliente") or "").strip()
    cotizacion.cliente_contacto = cliente_contacto or (cliente_sel.nombre if cliente_sel else cotizacion.cliente)
    cotizacion.cliente_razon_social = (
        cliente_sel.razon_social if cliente_sel else (request.form.get("cliente_razon_social") or "").strip()
    )
    cotizacion.cliente_cuit = cliente_sel.cuit if cliente_sel else (request.form.get("cliente_cuit") or "").strip()
    familia_form = normalizar_familia_cotizacion(request.form.get("familia"))
    cotizacion.familia = familia_form or familia_anterior
    cotizacion.moneda = moneda
    cotizacion.tipo_cambio_usado = tipo_cambio_usado
    cotizacion.condicion_iva = condicion_iva
    cotizacion.forma_pago = forma_pago
    cotizacion.condicion_cotizacion = condicion_cotizacion
    cotizacion.observacion_cliente = observacion_cliente
    cotizacion.carga_fiscal_pct = 0.0
    cotizacion.carga_fiscal_monto = 0.0
    preparar_seguimiento_cotizacion(cotizacion, cliente_sel=cliente_sel)

    descs = request.form.getlist("desc[]")
    detalles = request.form.getlist("detalle[]")
    row_ids = request.form.getlist("row_id[]")
    item_ids = request.form.getlist("item_id[]")
    imagenes_actuales = request.form.getlist("imagen_actual[]")
    cants = request.form.getlist("cant[]")
    costs = request.form.getlist("costo[]")
    iva_compras = request.form.getlist("iva_compra[]")
    extras = request.form.getlist("extra[]")
    margs = request.form.getlist("margen[]")
    descuentos = request.form.getlist("descuento[]")
    cargas_fiscales = request.form.getlist("carga_fiscal[]")
    iva_items = request.form.getlist("iva_item[]")

    items_existentes = {str(item.id): item for item in cotizacion.items if item.id}
    neto_total = 0.0
    iva_total = 0.0
    carga_fiscal_total_acum = 0.0
    imagenes_preparadas = {}

    try:
        for row_id in row_ids:
            file_storage = request.files.get(f"foto_{row_id}")
            imagen_preparada = preparar_imagen_producto(file_storage)
            if imagen_preparada:
                imagenes_preparadas[row_id] = imagen_preparada
    except ValueError as exc:
        flash(str(exc), "danger")
        destino = "editar_cotizacion" if es_edicion else "index"
        kwargs = {"id": cotizacion.id} if es_edicion else {}
        return None, redirect(url_for(destino, **kwargs))

    for i, desc in enumerate(descs):
        desc = (desc or "").strip()
        detalle = (detalles[i] if i < len(detalles) else "").strip()
        if not desc:
            continue

        costo = parsear_decimal(costs[i] if i < len(costs) else 0)
        iva_compra_pct = parsear_decimal(iva_compras[i] if i < len(iva_compras) else 0)
        extra_pct = parsear_decimal(extras[i] if i < len(extras) else 5.0, default=5.0)
        margen_pct = parsear_decimal(margs[i] if i < len(margs) else 0)
        descuento_pct = parsear_decimal(descuentos[i] if i < len(descuentos) else 0)
        carga_pct = parsear_decimal(cargas_fiscales[i] if i < len(cargas_fiscales) else 0)

        cantidad = normalizar_cantidad_entera(cants[i] if i < len(cants) else 1, default=1)
        costo = max(0.0, costo)
        iva_compra_pct = max(0.0, iva_compra_pct)
        extra_pct = max(0.0, extra_pct)
        margen_pct = max(0.0, margen_pct)
        descuento_pct = min(100.0, max(0.0, descuento_pct))
        carga_pct = max(0.0, carga_pct)
        margen_ratio = margen_pct / 100.0
        descuento_ratio = descuento_pct / 100.0

        iva_pct = normalizar_iva_venta(iva_items[i] if i < len(iva_items) else 21.0)

        costo_con_extra = costo * (1 + (extra_pct / 100.0))
        p_venta_lista = costo_con_extra * (1 + margen_ratio)
        p_venta_neto = p_venta_lista * (1 - descuento_ratio)
        sub_neto = cantidad * p_venta_neto
        monto_carga_item = sub_neto * (carga_pct / 100.0)
        iva_pct_aplicado = iva_pct
        monto_iva_item = sub_neto * (iva_pct_aplicado / 100.0)

        row_id = row_ids[i] if i < len(row_ids) else str(i)
        item_id = (item_ids[i] if i < len(item_ids) else "").strip()
        imagen_actual = (imagenes_actuales[i] if i < len(imagenes_actuales) else "").strip()
        imagen_local = guardar_imagen_producto(
            request.files.get(f"foto_{row_id}"), desc, row_id, optimized_bytes=imagenes_preparadas.get(row_id)
        )

        item = items_existentes.pop(item_id, None) if item_id else None
        if item and not imagen_actual:
            imagen_actual = item.imagen_url or ""

        imagen_final = imagen_local or imagen_actual or None
        if item and imagen_local and item.imagen_url and item.imagen_url != imagen_local:
            eliminar_imagen_local(item.imagen_url)

        if not item:
            item = ItemCotizacion()
            cotizacion.items.append(item)

        item.descripcion = desc
        item.detalle = detalle
        item.cantidad = cantidad
        item.costo_unitario = costo
        item.iva_compra_pct = iva_compra_pct
        item.costo_extra = extra_pct
        item.margen = margen_ratio
        item.descuento_pct = descuento_pct
        item.carga_fiscal = carga_pct
        item.iva_item = iva_pct_aplicado
        item.precio_venta = round(p_venta_neto, 2)
        item.subtotal = round(sub_neto + monto_iva_item, 2)
        item.imagen_url = imagen_final

        neto_total += sub_neto
        carga_fiscal_total_acum += monto_carga_item
        iva_total += monto_iva_item

    for item_sobrante in items_existentes.values():
        eliminar_imagen_local(item_sobrante.imagen_url)
        db.session.delete(item_sobrante)

    if not cotizacion.familia:
        flash("Debes definir la familia de la cotizacion antes de guardarla.", "danger")
        destino = "editar_cotizacion" if es_edicion else "index"
        kwargs = {"id": cotizacion.id} if es_edicion else {}
        return None, redirect(url_for(destino, **kwargs))

    if not cotizacion.cliente or not cotizacion.items:
        destino = "editar_cotizacion" if es_edicion else "index"
        kwargs = {"id": cotizacion.id} if es_edicion else {}
        return None, redirect(url_for(destino, **kwargs))

    total_bruto = neto_total + iva_total
    bonificacion_cierre_aplicada = min(bonificacion_cierre_solicitada, total_bruto)

    cotizacion.total_neto = round(neto_total, 2)
    cotizacion.total_iva = round(iva_total, 2)
    cotizacion.total_final = round(total_bruto - bonificacion_cierre_aplicada, 2)
    cotizacion.total_carga_fiscal = round(carga_fiscal_total_acum, 2)
    cotizacion.bonificacion_cierre_monto = round(bonificacion_cierre_aplicada, 2)

    if not cotizacion.numero_cotizacion:
        cotizacion.numero_cotizacion = generar_numero_cotizacion(cotizacion.fecha)
    db.session.add(cotizacion)
    db.session.flush()
    numero_ref = cotizacion.numero_cotizacion or str(cotizacion.id)
    registrar_auditoria(
        "Creó cotización" if not es_edicion else "Modificó cotización",
        "Cotización",
        entidad_id=cotizacion.id,
        entidad_ref=numero_ref,
        detalle=(
            f"Cliente: {cotizacion.cliente}. Familia: {cotizacion.familia}. Estado: {cotizacion.estado}. "
            f"Total: {cotizacion.moneda or 'ARS'} {cotizacion.total_final:,.2f}."
        ),
    )
    if es_edicion and estado_anterior and estado_anterior != cotizacion.estado:
        registrar_auditoria(
            "Cambió estado de cotización",
            "Cotización",
            entidad_id=cotizacion.id,
            entidad_ref=numero_ref,
            detalle=f"Estado anterior: {estado_anterior}. Estado nuevo: {cotizacion.estado}.",
        )
    db.session.commit()
    return cotizacion, None


@app.route("/api/tipo-cambio/oficial")
@token_required
def api_tipo_cambio_oficial():
    force = str(request.args.get("refresh") or "").strip().lower() in ("1", "true", "yes")
    solicitar_actualizacion_tipo_cambio_bna(force=force)
    payload = obtener_tipo_cambio_bna_cache()
    refrescando, ultimo_error = estado_actualizacion_tipo_cambio_bna()

    if payload:
        payload["refreshing"] = refrescando
        return jsonify(payload)
    if refrescando:
        return jsonify({"ok": True, "pending": True, "refreshing": True, "rate": None}), 202
    return jsonify({"ok": False, "error": ultimo_error or "No hay un tipo de cambio disponible todavia."}), 502


@app.route("/login", methods=["GET", "POST"])
def login():
    token = request.cookies.get("x-access-token")
    if token and obtener_usuario_desde_token(token):
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = buscar_usuario_por_username(username)
        if user and user.check_password(password):
            token = generar_token_usuario(user)
            destino = request.args.get("next") or request.form.get("next") or url_for("index")
            if not str(destino).startswith("/"):
                destino = url_for("index")
            res = redirect(destino)
            res.set_cookie("x-access-token", token, **auth_cookie_kwargs())
            return res
        error = "Usuario o contraseña incorrectos."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    res = redirect(url_for("login"))
    res.delete_cookie("x-access-token")
    return res


@app.route("/usuarios", methods=["GET", "POST"])
@token_required
@admin_required
def usuarios_page():
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        nombre_completo = (request.form.get("nombre_completo") or "").strip()
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""
        is_admin = request.form.get("is_admin") == "1"

        if not username:
            error = "El usuario es obligatorio."
        elif buscar_usuario_por_username(username):
            error = "Ese usuario ya existe."
        elif not password:
            error = "La contraseña es obligatoria."
        elif password != password_confirm:
            error = "Las contraseñas no coinciden."
        elif len(password) < 8:
            error = "La contraseña debe tener al menos 8 caracteres."
        else:
            nuevo = Usuario(username=username, nombre_completo=nombre_completo or None, is_admin=is_admin)
            nuevo.set_password(password)
            db.session.add(nuevo)
            db.session.flush()
            registrar_auditoria(
                "Creo usuario",
                "Usuario",
                entidad_id=nuevo.id,
                entidad_ref=nuevo.username,
                detalle=f"Rol asignado: {'Administrador' if nuevo.is_admin else 'Operador'}.",
            )
            db.session.commit()
            flash(f"Usuario {nuevo.username} creado correctamente.", "success")
            return redirect(url_for("usuarios_page"))

    usuarios = Usuario.query.order_by(Usuario.is_admin.desc(), func.lower(Usuario.username).asc()).all()
    total_usuarios = len(usuarios)
    total_admins = sum(1 for usuario in usuarios if usuario.is_admin)
    return render_template(
        "usuarios.html",
        error=error,
        usuarios=usuarios,
        total_usuarios=total_usuarios,
        total_admins=total_admins,
    )


@app.route("/usuarios/<int:id>/nombre", methods=["POST"])
@token_required
@admin_required
def actualizar_nombre_usuario(id):
    usuario = db.get_or_404(Usuario, id)
    nombre_completo = (request.form.get("nombre_completo") or "").strip()
    usuario.nombre_completo = nombre_completo or None
    registrar_auditoria(
        "Actualizo nombre de usuario",
        "Usuario",
        entidad_id=usuario.id,
        entidad_ref=usuario.username,
        detalle=f"Nombre visible para documentos: {usuario.nombre_para_documentos}.",
    )
    db.session.commit()
    flash(f"Nombre visible de {usuario.username} actualizado.", "success")
    return redirect(url_for("usuarios_page"))


@app.route("/familias", methods=["GET", "POST"])
@token_required
def familias_page():
    if request.method == "POST":
        nombre = normalizar_nombre_familia(request.form.get("nombre"))
        if not nombre:
            flash("El nombre de la familia es obligatorio.", "danger")
            return redirect(url_for("familias_page"))

        familia_existente = buscar_familia_por_nombre(nombre)
        if familia_existente:
            if not familia_existente.activa:
                familia_existente.activa = True
                registrar_auditoria(
                    "Reactivó familia",
                    "Familia",
                    entidad_id=familia_existente.id,
                    entidad_ref=familia_existente.nombre,
                )
                db.session.commit()
                flash(f"Familia {familia_existente.nombre} reactivada.", "success")
            else:
                flash("Esa familia ya existe.", "warning")
            return redirect(url_for("familias_page"))

        nueva = FamiliaCotizacion(nombre=nombre, activa=True)
        db.session.add(nueva)
        db.session.flush()
        registrar_auditoria("Creó familia", "Familia", entidad_id=nueva.id, entidad_ref=nueva.nombre)
        db.session.commit()
        flash(f"Familia {nueva.nombre} creada correctamente.", "success")
        return redirect(url_for("familias_page"))

    familias = FamiliaCotizacion.query.order_by(FamiliaCotizacion.activa.desc(), func.lower(FamiliaCotizacion.nombre)).all()
    usos = dict(db.session.query(Cotizacion.familia, func.count(Cotizacion.id)).group_by(Cotizacion.familia).all())
    return render_template("familias.html", familias=familias, usos_familias=usos)


@app.route("/familias/<int:id>/editar", methods=["POST"])
@token_required
def editar_familia(id):
    familia = db.get_or_404(FamiliaCotizacion, id)
    nombre_anterior = familia.nombre
    nombre_nuevo = normalizar_nombre_familia(request.form.get("nombre"))
    activa = request.form.get("activa") == "1"

    if not nombre_nuevo:
        flash("El nombre de la familia es obligatorio.", "danger")
        return redirect(url_for("familias_page"))

    existente = buscar_familia_por_nombre(nombre_nuevo)
    if existente and existente.id != familia.id:
        flash("Ya existe otra familia con ese nombre.", "danger")
        return redirect(url_for("familias_page"))

    familia.nombre = nombre_nuevo
    familia.activa = activa
    actualizadas = 0
    if nombre_anterior != nombre_nuevo:
        actualizadas = Cotizacion.query.filter(Cotizacion.familia == nombre_anterior).update(
            {Cotizacion.familia: nombre_nuevo}, synchronize_session=False
        )
    registrar_auditoria(
        "Modificó familia",
        "Familia",
        entidad_id=familia.id,
        entidad_ref=familia.nombre,
        detalle=f"Nombre anterior: {nombre_anterior}. Activa: {'si' if familia.activa else 'no'}. Cotizaciones actualizadas: {actualizadas}.",
    )
    db.session.commit()
    flash("Familia actualizada correctamente.", "success")
    return redirect(url_for("familias_page"))


@app.route("/familias/<int:id>/eliminar", methods=["POST"])
@token_required
def eliminar_familia(id):
    familia = db.get_or_404(FamiliaCotizacion, id)
    uso = Cotizacion.query.filter(Cotizacion.familia == familia.nombre).count()
    nombre_ref = familia.nombre
    if uso:
        familia.activa = False
        registrar_auditoria(
            "Desactivó familia",
            "Familia",
            entidad_id=familia.id,
            entidad_ref=nombre_ref,
            detalle=f"No se borro fisicamente porque tiene {uso} cotizaciones asociadas.",
        )
        db.session.commit()
        flash(f"Familia {nombre_ref} desactivada. Se mantiene en cotizaciones historicas.", "warning")
    else:
        db.session.delete(familia)
        registrar_auditoria("Eliminó familia", "Familia", entidad_id=id, entidad_ref=nombre_ref)
        db.session.commit()
        flash(f"Familia {nombre_ref} eliminada.", "success")
    return redirect(url_for("familias_page"))


@app.route("/api/familias", methods=["POST"])
@token_required
def agregar_familia_api():
    data = request.get_json(silent=True) or {}
    nombre = normalizar_nombre_familia(data.get("nombre"))
    if not nombre:
        return jsonify({"error": "Nombre de familia requerido."}), 400

    familia = buscar_familia_por_nombre(nombre)
    if familia:
        if not familia.activa:
            familia.activa = True
            registrar_auditoria("Reactivó familia", "Familia", entidad_id=familia.id, entidad_ref=familia.nombre)
            db.session.commit()
        return jsonify({"id": familia.id, "nombre": familia.nombre, "activa": bool(familia.activa)}), 200

    familia = FamiliaCotizacion(nombre=nombre, activa=True)
    db.session.add(familia)
    db.session.flush()
    registrar_auditoria("Creó familia", "Familia", entidad_id=familia.id, entidad_ref=familia.nombre)
    db.session.commit()
    return jsonify({"id": familia.id, "nombre": familia.nombre, "activa": True}), 201


@app.route("/api/clientes", methods=["POST"])
@token_required
def agregar_cliente():
    data = request.get_json(silent=True) or {}
    payload, error = validar_payload_cliente(data)
    if error:
        return jsonify({"error": error}), 400

    nuevo = Cliente(**payload)
    db.session.add(nuevo)
    db.session.commit()
    return jsonify(
        {
            "id": nuevo.id,
            "nombre": nuevo.nombre,
            "razon_social": nuevo.razon_social,
            "cuit": nuevo.cuit,
            "domicilio": nuevo.domicilio,
            "sector": nuevo.sector,
            "subsector": nuevo.subsector,
            "email": nuevo.email,
            "telefono": nuevo.telefono,
            "condicion_iva": nuevo.condicion_iva,
        }
    ), 201


@app.route("/api/clientes/<int:id>", methods=["PUT"])
@token_required
def actualizar_cliente(id):
    cliente = db.get_or_404(Cliente, id)
    data = request.get_json(silent=True) or {}
    payload, error = validar_payload_cliente(data, cliente_actual=cliente)
    if error:
        return jsonify({"error": error}), 400

    for campo, valor in payload.items():
        setattr(cliente, campo, valor)

    db.session.add(cliente)
    db.session.commit()
    return jsonify(
        {
            "id": cliente.id,
            "nombre": cliente.nombre,
            "razon_social": cliente.razon_social,
            "cuit": cliente.cuit,
            "domicilio": cliente.domicilio,
            "sector": cliente.sector,
            "subsector": cliente.subsector,
            "email": cliente.email,
            "telefono": cliente.telefono,
            "condicion_iva": cliente.condicion_iva,
        }
    ), 200


@app.route("/api/integracion/clientes", methods=["GET"])
@integration_token_required
def listar_clientes_integracion():
    q = (request.args.get("q") or "").strip()
    limit = min(parsear_entero_positivo(request.args.get("limit"), default=50) or 50, 200)

    query = Cliente.query
    if q:
        patron = f"%{q}%"
        query = query.filter(
            or_(
                Cliente.nombre.ilike(patron),
                Cliente.razon_social.ilike(patron),
                Cliente.cuit.ilike(patron),
                Cliente.email.ilike(patron),
                Cliente.telefono.ilike(patron),
            )
        )

    items = query.order_by(func.lower(Cliente.nombre).asc()).limit(limit).all()
    return jsonify({"items": [serializar_cliente(cliente) for cliente in items]})


@app.route("/api/integracion/clientes", methods=["POST"])
@integration_token_required
def agregar_cliente_integracion():
    data = request.get_json(silent=True) or {}
    payload, error = validar_payload_cliente(data)
    if error:
        return jsonify({"error": error}), 400

    nuevo = Cliente(**payload)
    db.session.add(nuevo)
    db.session.commit()
    return jsonify(serializar_cliente(nuevo)), 201


@app.route("/api/integracion/buscar")
@integration_token_required
def buscar_integracion():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"error": "q_required"}), 400

    patron = f"%{q}%"
    clientes = (
        Cliente.query.filter(
            or_(
                Cliente.nombre.ilike(patron),
                Cliente.razon_social.ilike(patron),
                Cliente.cuit.ilike(patron),
                Cliente.email.ilike(patron),
                Cliente.telefono.ilike(patron),
            )
        )
        .order_by(func.lower(Cliente.nombre).asc())
        .limit(10)
        .all()
    )
    cotizaciones = (
        Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
        .filter(
            or_(
                Cotizacion.numero_cotizacion.ilike(patron),
                Cotizacion.cliente.ilike(patron),
                Cotizacion.cliente_razon_social.ilike(patron),
                Cotizacion.cliente_cuit.ilike(patron),
                Cliente.nombre.ilike(patron),
                Cliente.razon_social.ilike(patron),
                Cliente.cuit.ilike(patron),
            )
        )
        .order_by(Cotizacion.fecha.desc(), Cotizacion.id.desc())
        .limit(10)
        .all()
    )

    return jsonify(
        {
            "query": q,
            "clientes": [serializar_cliente(cliente) for cliente in clientes],
            "cotizaciones": [serializar_cotizacion(cotizacion) for cotizacion in cotizaciones],
        }
    )


@app.route("/api/integracion/cotizaciones")
@integration_token_required
def listar_cotizaciones_integracion():
    q = (request.args.get("q") or "").strip()
    estado = normalizar_estado_cotizacion(request.args.get("estado"))
    cliente_id = parsear_entero_positivo(request.args.get("cliente_id"))
    limit = min(parsear_entero_positivo(request.args.get("limit"), default=50) or 50, 200)

    query = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    if q:
        patron = f"%{q}%"
        query = query.filter(
            or_(
                Cotizacion.numero_cotizacion.ilike(patron),
                Cotizacion.cliente.ilike(patron),
                Cotizacion.cliente_razon_social.ilike(patron),
                Cotizacion.cliente_cuit.ilike(patron),
                Cliente.nombre.ilike(patron),
                Cliente.razon_social.ilike(patron),
                Cliente.cuit.ilike(patron),
            )
        )
    if estado:
        query = query.filter(Cotizacion.estado == estado)
    if cliente_id:
        query = query.filter(Cotizacion.cliente_id == cliente_id)

    items = query.order_by(Cotizacion.fecha.desc(), Cotizacion.id.desc()).limit(limit).all()
    return jsonify({"items": [serializar_cotizacion(cotizacion) for cotizacion in items]})


@app.route("/api/integracion/cotizaciones/<int:id>")
@integration_token_required
def detalle_cotizacion_integracion(id):
    cotizacion = db.get_or_404(Cotizacion, id)
    return jsonify(serializar_cotizacion(cotizacion, incluir_items=True))


@app.before_request
def ensure_followup_worker():
    if app.config.get("TESTING"):
        return
    if app.debug and os.environ.get("WERKZEUG_RUN_MAIN") not in (None, "true"):
        return
    iniciar_worker_recordatorios()


@app.route("/cotizacion/<int:id>/estado", methods=["POST"])
@token_required
def actualizar_estado_cotizacion(id):
    cotizacion = db.get_or_404(Cotizacion, id)
    data = request.get_json(silent=True) or {}
    estado = normalizar_estado_cotizacion(data.get("estado"))
    if not estado:
        return jsonify({"error": "Estado invalido"}), 400
    estado_anterior = normalizar_estado_cotizacion(cotizacion.estado) or "En progreso"
    if estado == estado_anterior:
        return jsonify({"id": cotizacion.id, "estado": cotizacion.estado}), 200

    cotizacion.estado = estado
    registrar_auditoria(
        "Cambió estado de cotización",
        "Cotización",
        entidad_id=cotizacion.id,
        entidad_ref=cotizacion.numero_cotizacion or str(cotizacion.id),
        detalle=f"Estado anterior: {estado_anterior}. Estado nuevo: {estado}.",
    )
    db.session.commit()
    return jsonify({"id": cotizacion.id, "estado": cotizacion.estado}), 200


@app.route("/", methods=["GET", "POST"])
@token_required
def index():
    if request.method == "POST":
        _, redirect_response = persistir_cotizacion_desde_form()
        if redirect_response:
            return redirect_response
        return redirect(url_for("historial_page"))

    return renderizar_cotizador()


@app.route("/cotizacion/<int:id>/editar", methods=["GET", "POST"])
@token_required
def editar_cotizacion(id):
    cotizacion = db.get_or_404(Cotizacion, id)
    if request.method == "POST":
        _, redirect_response = persistir_cotizacion_desde_form(cotizacion)
        if redirect_response:
            return redirect_response
        return redirect(url_for("historial_page"))

    return renderizar_cotizador(cotizacion)


@app.route("/historial")
@token_required
def historial_page():
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    return render_template("historial.html", clientes=clientes)


@app.route("/dashboard")
@token_required
@admin_required
def dashboard_page():
    filtros = resolver_filtros_dashboard_request()
    estado = filtros["estado"]
    familia = filtros["familia"]
    sector = filtros["sector"]
    subsector = filtros["subsector"]
    moneda = filtros["moneda"]
    periodo = filtros["periodo"]
    desde = filtros["desde"]
    hasta = filtros["hasta"]
    desde_date = filtros["desde_date"]
    hasta_date = filtros["hasta_date"]

    query_periodo = construir_query_dashboard_periodo(filtros)
    contexto_operativo = construir_contexto_dashboard_operativo(query_periodo, filtros)

    series = construir_series_dashboard_sql(query_periodo, desde_date, hasta_date)
    agregado_periodo = series.pop("resumen")
    total = agregado_periodo["total"]
    aceptadas = agregado_periodo["aceptadas"]
    rechazadas = agregado_periodo["rechazadas"]
    en_progreso = agregado_periodo["en_progreso"]
    cerradas = aceptadas + rechazadas
    total_importe = agregado_periodo["total_importe"]
    importe_pipeline = agregado_periodo["importe_pipeline"]
    importe_aceptado = agregado_periodo["importe_aceptado"]
    ticket_promedio = round(total_importe / total, 2) if total else 0.0
    tasa_cierre = round((cerradas / total) * 100.0, 1) if total else 0.0
    tasa_aceptacion = round((aceptadas / total) * 100.0, 1) if total else 0.0

    dias_periodo = max((hasta_date - desde_date).days + 1, 1)
    desde_anterior = desde_date - timedelta(days=dias_periodo)
    hasta_anterior = desde_date - timedelta(days=1)
    query_previo = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    query_previo = aplicar_filtro_fecha_cotizaciones(query_previo, desde_anterior.isoformat(), hasta_anterior.isoformat())
    query_previo = aplicar_filtros_segmentacion_dashboard(
        query_previo, familia=familia, sector=sector, subsector=subsector, moneda=moneda
    )
    agregado_previo = construir_resumen_dashboard_sql(query_previo)
    total_previo = agregado_previo["total"]
    aceptadas_previas = agregado_previo["aceptadas"]

    resumen = {
        "total": total,
        "cerradas": cerradas,
        "aceptadas": aceptadas,
        "rechazadas": rechazadas,
        "en_progreso": en_progreso,
        "total_importe": total_importe,
        "importe_pipeline": importe_pipeline,
        "importe_aceptado": importe_aceptado,
        "ticket_promedio": ticket_promedio,
        "tasa_cierre": tasa_cierre,
        "tasa_aceptacion": tasa_aceptacion,
        "promedio_diario": round(total / dias_periodo, 1),
        "delta_total": calcular_variacion_porcentual(total, total_previo),
        "delta_aceptadas": calcular_variacion_porcentual(aceptadas, aceptadas_previas),
    }

    agrupaciones = construir_agrupaciones_dashboard_sql(query_periodo, total)
    top_clientes = agrupaciones["clientes"]
    familias_breakdown = agrupaciones["familias"]
    sectores_breakdown = agrupaciones["sectores"]
    subsectores_breakdown = agrupaciones["subsectores"]
    estado_foco_label = "Vista general del periodo"
    filtros_activos = []
    if familia:
        filtros_activos.append({"icon": "bi-diagram-3", "label": f"Familia: {familia}"})
    if sector:
        filtros_activos.append({"icon": "bi-building", "label": f"Sector: {sector}"})
    if subsector:
        filtros_activos.append({"icon": "bi-tags", "label": f"Subsector: {subsector}"})
    if filtros["cliente"]:
        filtros_activos.append({"icon": "bi-person-vcard", "label": f"Cliente: {filtros['cliente']}"})

    top_familia = familias_breakdown[0]["nombre"] if familias_breakdown else "Sin datos"
    top_sector = sectores_breakdown[0]["nombre"] if sectores_breakdown else "Sin datos"
    top_subsector = subsectores_breakdown[0]["nombre"] if subsectores_breakdown else "Sin datos"

    return render_template(
        "dashboard.html",
        resumen=resumen,
        series=series,
        top_clientes=top_clientes,
        familias_breakdown=familias_breakdown,
        sectores_breakdown=sectores_breakdown,
        subsectores_breakdown=subsectores_breakdown,
        estado_foco_label=estado_foco_label,
        filtro_desde_legible=desde_date.strftime("%d/%m/%Y"),
        filtro_hasta_legible=hasta_date.strftime("%d/%m/%Y"),
        dias_periodo=dias_periodo,
        filtros_activos=filtros_activos,
        familias_disponibles=obtener_familias_para_filtros(),
        sectores_cliente=SECTORES_CLIENTE,
        top_familia=top_familia,
        top_sector=top_sector,
        top_subsector=top_subsector,
        **contexto_operativo,
    )


@app.route("/dashboard/detalle-operativo")
@token_required
@admin_required
def dashboard_detalle_operativo():
    filtros = resolver_filtros_dashboard_request()
    query_periodo = construir_query_dashboard_periodo(filtros)
    contexto_operativo = construir_contexto_dashboard_operativo(query_periodo, filtros)
    return render_template(
        "_dashboard_operativo.html",
        familias_disponibles=obtener_familias_para_filtros(),
        **contexto_operativo,
    )


@app.route("/auditoria")
@token_required
@admin_required
def auditoria_page():
    usuario_filtro = (request.args.get("usuario") or "").strip()
    page = parsear_entero_positivo(request.args.get("page"), default=1) or 1
    query = Auditoria.query
    if usuario_filtro:
        query = query.filter(Auditoria.username == usuario_filtro)
    paginacion = paginar_query(
        query.order_by(Auditoria.fecha.desc(), Auditoria.id.desc()),
        page=page,
        per_page=AUDITORIA_PER_PAGE,
    )
    usuarios_auditoria = [row[0] for row in db.session.query(Auditoria.username).distinct().order_by(Auditoria.username.asc()).all()]
    return render_template(
        "auditoria.html",
        registros=paginacion["items"],
        usuario_filtro=usuario_filtro,
        usuarios_auditoria=usuarios_auditoria,
        paginacion={
            **paginacion,
            "pages_visible": construir_paginas_visibles(paginacion["page"], paginacion["pages"]),
        },
    )


@app.route("/filtrar_historial")
@token_required
def filtrar_historial():
    cliente_id_raw = (request.args.get("cliente_id") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    moneda = (request.args.get("moneda") or "").strip().upper()
    page = parsear_entero_positivo(request.args.get("page"), default=1) or 1

    query = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    if cliente_id_raw.isdigit():
        cliente_sel = db.session.get(Cliente, int(cliente_id_raw))
        if cliente_sel:
            condiciones_cliente = [Cotizacion.cliente_id == cliente_sel.id]
            if cliente_sel.nombre:
                condiciones_cliente.append(Cotizacion.cliente.ilike(f"%{cliente_sel.nombre}%"))
                condiciones_cliente.append(Cliente.nombre.ilike(f"%{cliente_sel.nombre}%"))
            if cliente_sel.razon_social:
                condiciones_cliente.append(Cotizacion.cliente_razon_social.ilike(f"%{cliente_sel.razon_social}%"))
                condiciones_cliente.append(Cliente.razon_social.ilike(f"%{cliente_sel.razon_social}%"))
            if cliente_sel.cuit:
                condiciones_cliente.append(Cotizacion.cliente_cuit == cliente_sel.cuit)
                condiciones_cliente.append(Cliente.cuit == cliente_sel.cuit)
            query = query.filter(or_(*condiciones_cliente))
        else:
            query = query.filter(Cotizacion.id == -1)
    elif cliente:
        patron = f"%{cliente}%"
        query = query.filter(
            or_(
                Cotizacion.cliente.ilike(patron),
                Cotizacion.cliente_razon_social.ilike(patron),
                Cliente.nombre.ilike(patron),
                Cliente.razon_social.ilike(patron),
                Cliente.cuit.ilike(patron),
            )
        )
    query = aplicar_filtro_fecha_cotizaciones(query, desde, hasta)
    if moneda in ("ARS", "USD"):
        query = query.filter(Cotizacion.moneda == moneda)
    query = query.options(contains_eager(Cotizacion.cliente_ref))

    paginacion = paginar_query(query.order_by(Cotizacion.id.desc()), page=page, per_page=HISTORIAL_PER_PAGE)
    cotizaciones = paginacion["items"]
    resultados = [
        {
            "id": c.id,
            "numero_cotizacion": c.numero_cotizacion or str(c.id).zfill(4),
            "estado": normalizar_estado_cotizacion(c.estado) or "En progreso",
            "cliente": c.cliente,
            "cliente_nombre": (c.cliente_ref.nombre if c.cliente_ref else c.cliente) or "Sin Cliente",
            "familia": c.familia or "",
            "fecha": c.fecha.strftime("%d/%m/%Y"),
            "moneda": c.moneda or "ARS",
            "total_final": c.total_final or 0.0,
            "total": f"${c.total_final:,.2f}",
        }
        for c in cotizaciones
    ]
    return jsonify(
        {
            "items": resultados,
            "pagination": {
                **{k: paginacion[k] for k in ("page", "per_page", "total", "pages", "has_prev", "has_next", "prev_page", "next_page", "start", "end")},
                "pages_visible": construir_paginas_visibles(paginacion["page"], paginacion["pages"]),
            },
        }
    )


@app.route("/cotizacion/<int:id>")
@token_required
def ver_cotizacion(id):
    cot = db.get_or_404(Cotizacion, id)
    return render_template("cotizacion_cliente.html", **construir_contexto_documento_cotizacion(cot))


@app.route("/cotizacion/<int:id>/llave-en-mano")
@token_required
def ver_cotizacion_llave_en_mano(id):
    cot = db.get_or_404(Cotizacion, id)
    return render_template("cotizacion_llave_en_mano.html", **construir_contexto_llave_en_mano(cot))


@app.route("/cotizacion/<int:id>/xlsx")
@token_required
def exportar_cotizacion_xlsx(id):
    cotizacion = db.get_or_404(Cotizacion, id)
    contenido = generar_excel_cotizacion(cotizacion)
    nombre_base = (cotizacion.numero_cotizacion or f"cotizacion-{cotizacion.id}").replace("/", "-").replace("\\", "-")
    return Response(
        contenido,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_base}.xlsx"'},
    )


@app.route("/cotizacion/clonar/<int:id>", methods=["POST"])
@token_required
def clonar_cotizacion(id):
    cot_original = db.get_or_404(Cotizacion, id)
    fecha_clon = datetime.utcnow()
    numero_nuevo = generar_numero_cotizacion(fecha_clon)
    imagenes_clonadas = []

    nueva_cotizacion = Cotizacion(
        numero_cotizacion=numero_nuevo,
        estado="En progreso",
        seguimiento_activo=False,
        seguimiento_email=cot_original.seguimiento_email,
        seguimiento_cada_dias=cot_original.seguimiento_cada_dias,
        seguimiento_proximo_envio=None,
        seguimiento_ultimo_envio=None,
        nombre_fantasia=cot_original.nombre_fantasia or NOMBRE_FANTASIA,
        razon_social=cot_original.razon_social or RAZON_SOCIAL,
        cuit=cot_original.cuit or CUIT,
        cliente=cot_original.cliente,
        cliente_id=cot_original.cliente_id,
        cliente_contacto=cot_original.cliente_contacto,
        cliente_razon_social=cot_original.cliente_razon_social,
        cliente_cuit=cot_original.cliente_cuit,
        familia=cot_original.familia,
        fecha=fecha_clon,
        moneda=cot_original.moneda or "USD",
        tipo_cambio_usado=cot_original.tipo_cambio_usado,
        condicion_iva=cot_original.condicion_iva,
        condicion_cotizacion=cot_original.condicion_cotizacion,
        forma_pago=cot_original.forma_pago,
        observacion_cliente=cot_original.observacion_cliente,
        carga_fiscal_pct=cot_original.carga_fiscal_pct or 0.0,
        carga_fiscal_monto=cot_original.carga_fiscal_monto or 0.0,
        total_carga_fiscal=cot_original.total_carga_fiscal or 0.0,
        bonificacion_cierre_monto=cot_original.bonificacion_cierre_monto or 0.0,
        total_neto=cot_original.total_neto or 0.0,
        total_iva=cot_original.total_iva or 0.0,
        total_final=cot_original.total_final or 0.0,
    )

    db.session.add(nueva_cotizacion)
    db.session.flush()

    try:
        for item in cot_original.items:
            imagen_clonada = clonar_imagen_local(item.imagen_url)
            if imagen_clonada and imagen_clonada != item.imagen_url and es_ruta_imagen_local(imagen_clonada):
                imagenes_clonadas.append(imagen_clonada)

            nueva_cotizacion.items.append(
                ItemCotizacion(
                    descripcion=item.descripcion,
                    detalle=item.detalle,
                    cantidad=item.cantidad,
                    costo_unitario=item.costo_unitario,
                    iva_compra_pct=item.iva_compra_pct or 0.0,
                    costo_extra=item.costo_extra or 0.0,
                    margen=item.margen,
                    descuento_pct=item.descuento_pct or 0.0,
                    carga_fiscal=item.carga_fiscal or 0.0,
                    iva_item=item.iva_item or 0.0,
                    precio_venta=item.precio_venta,
                    subtotal=item.subtotal,
                    imagen_url=imagen_clonada,
                )
            )

        registrar_auditoria(
            "Clono cotizacion",
            "Cotizacion",
            entidad_id=nueva_cotizacion.id,
            entidad_ref=nueva_cotizacion.numero_cotizacion or str(nueva_cotizacion.id),
            detalle=(
                f"Origen: {cot_original.numero_cotizacion or cot_original.id}. "
                f"Cliente: {nueva_cotizacion.cliente or 'Sin cliente'}. "
                f"Nuevo total: {nueva_cotizacion.moneda or 'ARS'} {nueva_cotizacion.total_final or 0.0:,.2f}."
            ),
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        for ruta in imagenes_clonadas:
            eliminar_imagen_local(ruta)
        raise

    return jsonify(
        {
            "ok": True,
            "nueva_id": nueva_cotizacion.id,
            "redirect_url": url_for("editar_cotizacion", id=nueva_cotizacion.id),
            "mensaje": "Cotizacion duplicada con exito.",
        }
    )


@app.route("/cotizacion/<int:id>/eliminar", methods=["POST"])
@token_required
@admin_required
def eliminar_cotizacion(id):
    cotizacion = db.get_or_404(Cotizacion, id)
    numero_ref = cotizacion.numero_cotizacion or str(cotizacion.id)
    cliente_ref = cotizacion.cliente or "Sin cliente"
    familia_ref = cotizacion.familia or "Sin familia"
    moneda_ref = cotizacion.moneda or "ARS"
    total_ref = cotizacion.total_final or 0.0
    estado_ref = cotizacion.estado or "En progreso"

    for item in cotizacion.items:
        eliminar_imagen_local(item.imagen_url)

    db.session.delete(cotizacion)
    registrar_auditoria(
        "Eliminó cotización",
        "Cotización",
        entidad_id=id,
        entidad_ref=numero_ref,
        detalle=f"Cliente: {cliente_ref}. Familia: {familia_ref}. Estado: {estado_ref}. Total: {moneda_ref} {total_ref:,.2f}.",
    )
    db.session.commit()

    return jsonify({"ok": True, "id": id, "numero_cotizacion": numero_ref}), 200


if __name__ == "__main__":
    app.run(debug=True)
